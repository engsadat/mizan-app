"""Excel is the system of record this phase. SQLite is login/users only.

Column maps match the live map/org-chart readers (not the broken import_projects.py map).
"""
from __future__ import annotations

import os
import shutil
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace

import openpyxl

ERR_VALS = {'#REF!', '=#REF!', '#VALUE!', '#N/A', '#NAME?', '#DIV/0!', '#NULL!', '#NUM!'}
STATUS_ON_STRENGTH = 'على قوة العمل'
STATUS_REPLACEMENT = 'بديل'
ONGOING_STATE = 'تحت التنفيذ'

# employees data source.xlsx sheet `data` (0-based)
EMP_COL = {
    'phone': 3,
    'hire_date': 4,
    'id_number': 5,
    'email': 2,
    'contract_end': 9,
    'resignation': 10,
    'status': 11,
    'status_reason': 12,
    'nationality': 14,
    'kafala': 15,
    'region': 16,
    'cbu': 17,
    'category': 18,
    'office': 19,
    'name': 20,
    'job': 21,
    'serial': 22,
    'unit_price': 23,
    'manager': 24,
    'att_group': 25,
    're': 27,
}

# project_2026_database_ver1_updated.xlsx sheet `pro` (0-based)
# Verified against /reports/project-map-smart (132 ongoing).
PRO_COL = {
    'x': 3,
    'y': 4,
    'po': 9,
    'included': 10,
    'name': 12,
    'wad': 16,
    'region': 17,
    'contractor': 18,
    'state': 23,
    'start': 24,
    'end': 25,
    'ptype': 28,
    'value': 29,
}


def sv(v):
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ERR_VALS else s


def load_copy(path):
    path = Path(path)
    if not path.exists():
        return None
    tmp = str(path) + '.tmp.xlsx'
    shutil.copy2(str(path), tmp)
    try:
        return openpyxl.load_workbook(tmp, data_only=True)
    finally:
        if Path(tmp).exists():
            os.remove(tmp)


def _sources():
    try:
        from flask import current_app, has_app_context
        if has_app_context():
            return current_app.config.get('EXCEL_SOURCES') or {}
    except Exception:
        pass
    return {}


def _default_data_dir():
    return Path(__file__).resolve().parent.parent / 'data'


def employees_path():
    src = _sources()
    if src.get('employees'):
        return Path(src['employees'])
    return _default_data_dir() / 'source' / 'employees data source.xlsx'


def projects_path():
    src = _sources()
    if src.get('projects'):
        return Path(src['projects'])
    return _default_data_dir() / 'source' / 'project_2026_database_ver1_updated.xlsx'


def office_re_path():
    src = _sources()
    if src.get('office_re'):
        return Path(src['office_re'])
    return _default_data_dir() / 'Organize' / 'Office-RE.xlsx'


def norm_person(s):
    """Normalize Arabic person/office names for matching (spaces, alef, yeh)."""
    s = sv(s).replace('م/', ' ').replace('م /', ' ')
    s = ' '.join(s.split())
    return s.translate(str.maketrans({
        'ى': 'ي', 'ئ': 'ي', 'ؤ': 'و',
        'أ': 'ا', 'إ': 'ا', 'آ': 'ا',
        'ة': 'ه',
    }))


def match_office_name(raw, canon_by_norm):
    """Map an employee/project RE string to a canonical Office-RE name."""
    n = norm_person(raw)
    if not n:
        return None
    if n in canon_by_norm:
        return canon_by_norm[n]
    hits = [c for nn, c in canon_by_norm.items() if n in nn or nn in n]
    uniq = list(dict.fromkeys(hits))
    if len(uniq) == 1:
        return uniq[0]
    return None


def _cell(row, idx):
    return row[idx] if row and len(row) > idx else None


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = sv(v)
    if not s:
        return None
    for fmt, n in (
        ('%Y-%m-%d %H:%M:%S', 19),
        ('%Y-%m-%d', 10),
        ('%d/%m/%Y', 10),
        ('%d-%m-%Y', 10),
    ):
        try:
            return datetime.strptime(s[:n], fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(v):
    s = sv(v).replace(',', '')
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


class SimplePagination:
    def __init__(self, items, page, per_page, total):
        self.items = items
        self.page = page
        self.per_page = per_page
        self.total = total
        self.pages = (total + per_page - 1) // per_page if total else 0

    def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
        last = 0
        for num in range(1, self.pages + 1):
            if (
                num <= left_edge
                or (self.page - left_current - 1 < num < self.page + right_current)
                or num > self.pages - right_edge
            ):
                if last + 1 != num:
                    yield None
                yield num
                last = num


def _ns(**kw):
    return SimpleNamespace(**kw)


def _emp_from_row(row, row_idx):
    name = sv(_cell(row, EMP_COL['name']))
    if not name:
        return None
    job = sv(_cell(row, EMP_COL['job']))
    nat = sv(_cell(row, EMP_COL['nationality']))
    status = sv(_cell(row, EMP_COL['status']))
    office = sv(_cell(row, EMP_COL['office']))
    att = sv(_cell(row, EMP_COL['att_group']))
    cat = sv(_cell(row, EMP_COL['category']))
    return _ns(
        id=row_idx,
        full_name=name,
        phone=sv(_cell(row, EMP_COL['phone'])),
        email=sv(_cell(row, EMP_COL['email'])),
        id_number=sv(_cell(row, EMP_COL['id_number'])),
        hire_date=_parse_date(_cell(row, EMP_COL['hire_date'])),
        contract_end_date=_parse_date(_cell(row, EMP_COL['contract_end'])),
        resignation_date=_parse_date(_cell(row, EMP_COL['resignation'])),
        status_reason=sv(_cell(row, EMP_COL['status_reason'])),
        kafala=sv(_cell(row, EMP_COL['kafala'])),
        region=sv(_cell(row, EMP_COL['region'])),
        cbu=sv(_cell(row, EMP_COL['cbu'])),
        category=cat,
        supervision_type=cat,
        profession='',
        qualification='',
        grad_year=None,
        experience_years=None,
        serial_no=sv(_cell(row, EMP_COL['serial'])),
        re_code=sv(_cell(row, EMP_COL['re'])),
        direct_manager=sv(_cell(row, EMP_COL['manager'])),
        unit_price=_parse_float(_cell(row, EMP_COL['unit_price'])),
        is_replacement=(status == STATUS_REPLACEMENT),
        job_code=_ns(title=job, id=job) if job else None,
        nationality=_ns(name_ar=nat, id=nat) if nat else None,
        current_status=_ns(name_ar=status, id=status) if status else None,
        office=_ns(name=office, id=office) if office else None,
        attendance_group=_ns(name=att, id=att) if att else None,
        replaced_employee=None,
        replacement_start_date=None,
        replacement_end_date=None,
        status_history=[],
    )


def load_employees():
    path = employees_path()
    wb = load_copy(path)
    if wb is None:
        return []
    try:
        ws = wb['data']
    except KeyError:
        ws = wb.active
    out = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        emp = _emp_from_row(row, idx)
        if emp:
            out.append(emp)
    return out


def get_employee(emp_id):
    for emp in load_employees():
        if emp.id == emp_id:
            return emp
    return None


def employee_statuses(employees=None):
    employees = employees if employees is not None else load_employees()
    names = []
    seen = set()
    for e in employees:
        n = e.current_status.name_ar if e.current_status else ''
        if n and n not in seen:
            seen.add(n)
            names.append(n)
    return [_ns(id=n, name_ar=n) for n in names]


def load_projects(included_only=True):
    path = projects_path()
    wb = load_copy(path)
    if wb is None:
        return []
    try:
        ws = wb['pro']
    except KeyError:
        ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        inc = sv(_cell(row, PRO_COL['included'])).lower()
        if included_only and inc != 'yes':
            continue
        name = sv(_cell(row, PRO_COL['name'])) or 'بدون اسم'
        out.append(_ns(
            name=name,
            region=sv(_cell(row, PRO_COL['region'])),
            project_state=sv(_cell(row, PRO_COL['state'])),
            value=_parse_float(_cell(row, PRO_COL['value'])) or 0.0,
            included=(inc == 'yes'),
        ))
    return out


def load_office_re():
    """Rows from Office-RE.xlsx sheet RE_Mail."""
    wb = load_copy(office_re_path())
    if wb is None:
        return []
    try:
        ws = wb['RE_Mail']
    except KeyError:
        ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = sv(_cell(row, 1))
        region = sv(_cell(row, 4))
        if not name or not region:
            continue
        out.append(_ns(
            name=name,
            region=region,
            phone=sv(_cell(row, 3)),
            office=sv(_cell(row, 8)),
        ))
    return out
