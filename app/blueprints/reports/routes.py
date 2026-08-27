import json, base64, re
from datetime import date
from collections import Counter
from pathlib import Path
from flask import render_template, request
from flask_login import login_required
from app.blueprints.reports import reports_bp
from app.models import Employee, EmployeeStatus, JobCode, AttendanceGroup, Nationality, Office, Project
from app import db

REGIONS = ['عسير', 'جازان', 'الباحة', 'نجران']
REGION_CODES = {'عسير': 'AS', 'جازان': 'JZ', 'الباحة': 'BA', 'نجران': 'NJ'}
ONGOING_STATE = 'تحت التنفيذ'


def _b64_logo(rel_path):
    try:
        # Always look in app/static/img/ first (works locally + PythonAnywhere)
        static_path = Path(__file__).parent.parent.parent / 'static' / 'img' / rel_path
        if static_path.exists():
            return 'data:image/png;base64,' + base64.b64encode(static_path.read_bytes()).decode()
        # Fallback: HR project NWC layout folder (local dev only)
        hr_root = Path(__file__).parent.parent.parent.parent.parent.parent.parent
        fallback = hr_root / 'NWC layout' / 'img' / rel_path
        if fallback.exists():
            return 'data:image/png;base64,' + base64.b64encode(fallback.read_bytes()).decode()
    except Exception:
        pass
    return ''


def _chart_data():
    rows = (
        db.session.query(Employee.region, Employee.category, JobCode.title)
        .join(EmployeeStatus, Employee.current_status_id == EmployeeStatus.id)
        .outerjoin(JobCode, Employee.job_code_id == JobCode.id)
        .filter(EmployeeStatus.name_ar == 'على قوة العمل')
        .all()
    )

    def cnt(fn):
        return [sum(1 for r in rows if r.region == reg and fn(r)) for reg in REGIONS]

    def has(kws):
        return lambda r: any(k in (r.title or '') for k in kws)

    def is_job(kw):
        return lambda r: kw in (r.title or '')

    charts = [
        ('c1',  'إجمالي عناصر المشروع', 'h-special', cnt(lambda r: True)),
        ('c2',  'عناصر الإشراف',         'h-special', cnt(lambda r: (r.category or '') == 'إشراف')),
        ('c3',  'عناصر الدعم الفني',     'h-special', cnt(lambda r: (r.category or '') == 'دعم فني')),
        ('c4',  'مهندس مقيم',            'h-navy',    cnt(is_job('مهندس مقيم'))),
        ('c5',  'مهندس موقع',            'h-navy',    cnt(is_job('مهندس موقع'))),
        ('c6',  'مهندس ميكانيكا',         'h-navy',    cnt(has(['ميكانيك']))),
        ('c7',  'مهندس كهرباء وتحكم',    'h-navy',    cnt(has(['كهرباء', 'تحكم']))),
        ('c8',  'مهندس أمن وسلامة',      'h-navy',    cnt(lambda r: 'مهندس' in (r.title or '') and ('أمن' in (r.title or '') or 'سلامة' in (r.title or '')))),
        ('c9',  'مهندس تخطيط',           'h-navy',    cnt(has(['تخطيط']))),
        ('c10', 'مهندس مواد وجودة',      'h-navy',    cnt(lambda r: 'مواد' in (r.title or '') or ('جودة' in (r.title or '') and 'مهندس' in (r.title or '')))),
        ('c11', 'حاسب كميات',            'h-navy',    cnt(has(['كميات']))),
        ('c12', 'مراقب موقع',            'h-navy',    cnt(lambda r: 'مراقب موقع' in (r.title or ''))),
        ('c13', 'مساح',                  'h-navy',    cnt(has(['مساح']))),
        ('c14', 'مراقب أمن وسلامة',      'h-navy',    cnt(lambda r: 'مراقب' in (r.title or '') and ('أمن' in (r.title or '') or 'سلامة' in (r.title or '')))),
        ('c15', 'إخصائي وثائق',          'h-navy',    cnt(has(['وثائق']))),
    ]

    region_totals = {reg: sum(1 for r in rows if r.region == reg) for reg in REGIONS}
    total = len(rows)
    return charts, region_totals, total


@reports_bp.route('/')
@login_required
def index():
    return render_template('reports/index.html')


@reports_bp.route('/org-chart')
@login_required
def org_chart_selector():
    """Org chart region selector page."""
    regions = [
        {'code': 'عسير', 'name': 'عسير', 'file': '09_OrgChart_Asir.html'},
        {'code': 'جازان', 'name': 'جازان', 'file': '10_OrgChart_Jizan.html'},
        {'code': 'الباحة', 'name': 'الباحة', 'file': '11_OrgChart_Baha.html'},
        {'code': 'نجران', 'name': 'نجران', 'file': '12_OrgChart_Najran.html'},
    ]
    return render_template('reports/org_chart_selector.html', regions=regions)


@reports_bp.route('/org-chart/<region>')
@login_required
def org_chart_view(region):
    """Serve org chart HTML for a region."""
    valid_regions = {'عسير': '09_OrgChart_Asir.html', 'جازان': '10_OrgChart_Jizan.html',
                     'الباحة': '11_OrgChart_Baha.html', 'نجران': '12_OrgChart_Najran.html'}

    if region not in valid_regions:
        return render_template('error.html', error='Invalid region'), 404

    file_path = Path(__file__).parent.parent.parent / 'static' / 'org_charts' / valid_regions[region]
    if not file_path.exists():
        return render_template('error.html', error='Org chart not found'), 404

    html_content = file_path.read_text(encoding='utf-8')
    return html_content, 200, {'Content-Type': 'text/html; charset=utf-8'}


REPORT_COLUMNS = [
    ('job',        'الوظيفة',          True),
    ('region',     'المنطقة',          True),
    ('nationality','الجنسية',          True),
    ('supervision','نوع الإشراف',      True),
    ('status',     'الموقف',           False),
    ('phone',      'الهاتف',           False),
    ('office',     'المكتب',           False),
    ('att_group',  'كشف الحضور',       False),
    ('hire_date',  'تاريخ الالتحاق',  False),
    ('end_date',   'نهاية العقد',      False),
    ('kafala',     'الكفالة',          False),
    ('re_code',    'كود RE',           False),
]


@reports_bp.route('/filter-report', methods=['GET', 'POST'])
@login_required
def filter_report():
    job_codes    = JobCode.query.order_by(JobCode.title).all()
    att_groups   = AttendanceGroup.query.order_by(AttendanceGroup.name).all()
    all_statuses = EmployeeStatus.query.order_by(EmployeeStatus.name_ar).all()
    kafala_opts  = sorted(set(
        r[0] for r in db.session.query(Employee.kafala).filter(Employee.kafala.isnot(None), Employee.kafala != '').all()
    ))
    offices      = Office.query.order_by(Office.name).all()

    employees       = []
    filters_applied = False
    selected_cols   = [k for k, _, default in REPORT_COLUMNS if default]
    f               = {}

    if request.method == 'POST':
        filters_applied = True
        f = {
            'regions':      request.form.getlist('regions'),
            'nat_group':    request.form.get('nat_group', ''),
            'supervision':  request.form.get('supervision', ''),
            'status_ids':   request.form.getlist('status_ids'),
            'job_code_ids': request.form.getlist('job_code_ids'),
            'att_group_id': request.form.getlist('att_group_id'),
            'kafala_vals':  request.form.getlist('kafala_vals'),
            'office_ids':   request.form.getlist('office_ids'),
        }
        selected_cols = request.form.getlist('cols') or selected_cols

        q = Employee.query.outerjoin(Employee.current_status)
        if f['regions']:
            q = q.filter(Employee.region.in_(f['regions']))
        if f['nat_group'] == 'سعودي':
            q = q.join(Employee.nationality).filter(Nationality.name_ar == 'سعودي')
        elif f['nat_group'] == 'غير سعودي':
            q = q.join(Employee.nationality).filter(Nationality.name_ar != 'سعودي')
        if f['supervision']:
            q = q.filter(Employee.supervision_type == f['supervision'])
        if f['status_ids']:
            q = q.filter(Employee.current_status_id.in_([int(x) for x in f['status_ids']]))
        if f['job_code_ids']:
            q = q.filter(Employee.job_code_id.in_([int(x) for x in f['job_code_ids']]))
        if f['att_group_id']:
            q = q.filter(Employee.attendance_group_id.in_([int(x) for x in f['att_group_id']]))
        if f['kafala_vals']:
            q = q.filter(Employee.kafala.in_(f['kafala_vals']))
        if f['office_ids']:
            q = q.filter(Employee.office_id.in_([int(x) for x in f['office_ids']]))

        employees = q.order_by(Employee.region, Employee.full_name).all()

    return render_template('reports/filter_report.html',
                           job_codes=job_codes,
                           att_groups=att_groups,
                           all_statuses=all_statuses,
                           kafala_opts=kafala_opts,
                           offices=offices,
                           regions=REGIONS,
                           report_columns=REPORT_COLUMNS,
                           employees=employees,
                           filters_applied=filters_applied,
                           selected_cols=selected_cols,
                           f=f,
                           logo_nwc=_b64_logo('NWC_Logo.png'),
                           logo_alamro=_b64_logo('Alamro_Logo.png'))


@reports_bp.route('/emp-kpi')
@login_required
def emp_kpi():
    region = request.args.get('region', '')

    base_q = Employee.query.join(Employee.current_status).filter(
        EmployeeStatus.name_ar == 'على قوة العمل'
    )
    if region and region in REGIONS:
        base_q = base_q.filter(Employee.region == region)
    employees = base_q.all()

    total_active = len(employees)
    saudi_count  = sum(1 for e in employees if e.nationality and e.nationality.name_ar == 'سعودي')
    non_saudi    = total_active - saudi_count

    repl_q = Employee.query.join(Employee.current_status).filter(EmployeeStatus.name_ar == 'بديل')
    if region and region in REGIONS:
        repl_q = repl_q.filter(Employee.region == region)
    replacement_count = repl_q.count()

    # Region bars
    region_counts = Counter(e.region for e in employees if e.region)
    region_data   = [{'name': r, 'count': region_counts.get(r, 0)} for r in REGIONS]

    # Kafala bars
    kafala_counts = Counter(e.kafala for e in employees if e.kafala)
    kafala_data   = [{'name': k, 'count': v}
                     for k, v in sorted(kafala_counts.items(), key=lambda x: -x[1])]

    # Nationality top bars
    nat_counts = Counter(e.nationality.name_ar for e in employees if e.nationality)
    top_nats   = nat_counts.most_common(8)
    top_keys   = {n for n, _ in top_nats}
    other      = sum(v for k, v in nat_counts.items() if k not in top_keys)
    nat_data   = [{'name': n, 'count': c} for n, c in top_nats]
    if other:
        nat_data.append({'name': 'أخرى', 'count': other})

    # E2 / E3 table
    job_e = {}
    for emp in employees:
        if not emp.job_code:
            continue
        title = emp.job_code.title
        m = re.search(r'\b(E[23])\b', title)
        if not m:
            continue
        level = m.group(1)
        base  = title[:title.rfind(level)].strip()
        job_e.setdefault(base, {'E2': 0, 'E3': 0})
        job_e[base][level] += 1

    e_table = sorted(
        [{'job': b, 'e2': v['E2'], 'e3': v['E3'], 'total': v['E2'] + v['E3']}
         for b, v in job_e.items()],
        key=lambda x: -x['total']
    )

    from datetime import date as _date
    return render_template('reports/emp_kpi.html',
                           now=_date.today().strftime('%Y-%m-%d'),
                           total_active=total_active,
                           saudi_count=saudi_count,
                           non_saudi=non_saudi,
                           replacement_count=replacement_count,
                           region_data=json.dumps(region_data, ensure_ascii=False),
                           kafala_data=json.dumps(kafala_data, ensure_ascii=False),
                           nat_data=json.dumps(nat_data, ensure_ascii=False),
                           e_table=e_table,
                           selected_region=region,
                           regions=REGIONS,
                           logo_nwc=_b64_logo('NWC_Logo.png'),
                           logo_alamro=_b64_logo('Alamro_Logo.png'))


@reports_bp.route('/emp-dashboard')
@login_required
def emp_dashboard():
    charts, region_totals, total = _chart_data()
    charts_json = json.dumps([
        {'id': cid, 'title': title, 'cls': cls, 'data': data}
        for cid, title, cls, data in charts
    ], ensure_ascii=False)
    region_bar = [
        {'code': REGION_CODES[r], 'name': r, 'count': region_totals.get(r, 0)}
        for r in REGIONS
    ]
    logo_alamro = _b64_logo('Alamro_Logo.png')
    logo_nwc    = _b64_logo('NWC_Logo.png')
    return render_template(
        'reports/emp_dashboard.html',
        charts_json=charts_json,
        region_bar=region_bar,
        total=total,
        logo_alamro=logo_alamro,
        logo_nwc=logo_nwc,
    )

def _msar(v):
    return round((v or 0) / 1e6, 1)


def _project_pivot_data():
    """Included=Yes. KPIs + charts = تحت التنفيذ by region. Tables = all statuses."""
    projects = Project.query.filter_by(included=True).all()

    def region_of(p):
        return (p.region or '').strip() or 'غير محدد'

    def state_of(p):
        return (p.project_state or '').strip() or 'غير محدد'

    extra = []
    for p in projects:
        r = region_of(p)
        if r not in REGIONS and r not in extra:
            extra.append(r)
    regions = list(REGIONS) + extra

    status_value = {}
    pivot = {}
    for p in projects:
        st, rg = state_of(p), region_of(p)
        status_value[st] = status_value.get(st, 0) + (p.value or 0)
        pivot.setdefault(st, {})
        cell = pivot[st].setdefault(rg, {'count': 0, 'value': 0.0})
        cell['count'] += 1
        cell['value'] += p.value or 0

    statuses = sorted(status_value, key=lambda s: -status_value[s])
    for st in statuses:
        for rg in regions:
            cell = pivot[st].setdefault(rg, {'count': 0, 'value': 0.0})
            cell['msar'] = _msar(cell['value'])

    region_counts = {}
    region_msar = {}
    for rg in regions:
        tot = sum(pivot[st][rg]['count'] for st in statuses) if statuses else 0
        val = sum(pivot[st][rg]['value'] for st in statuses) if statuses else 0
        region_counts[rg] = tot
        region_msar[rg] = _msar(val)

    count_row_totals = {
        st: sum(pivot[st][rg]['count'] for rg in regions) for st in statuses
    }
    msar_row_totals = {
        st: _msar(sum(pivot[st][rg]['value'] for rg in regions)) for st in statuses
    }

    ongoing = [p for p in projects if state_of(p) == ONGOING_STATE]
    ongoing_kpis = []

    # Add total card first
    ongoing_kpis.append({
        'region': 'الإجمالي',
        'count': len(ongoing),
        'msar': _msar(sum(p.value or 0 for p in ongoing)),
        'is_total': True,
    })

    # Then add by region
    for rg in regions:
        rows = [p for p in ongoing if region_of(p) == rg]
        ongoing_kpis.append({
            'region': rg,
            'count': len(rows),
            'msar': _msar(sum(p.value or 0 for p in rows)),
            'is_total': False,
        })

    return {
        'regions': regions,
        'statuses': statuses,
        'pivot': pivot,
        'region_counts': region_counts,
        'region_msar': region_msar,
        'count_row_totals': count_row_totals,
        'msar_row_totals': msar_row_totals,
        'total_projects': len(projects),
        'total_msar': _msar(sum(p.value or 0 for p in projects)),
        'ongoing_kpis': ongoing_kpis,
        'ongoing_count': len(ongoing),
        'ongoing_msar': _msar(sum(p.value or 0 for p in ongoing)),
        'value_chart': [{'name': k['region'], 'value': k['msar']} for k in ongoing_kpis],
        'count_chart': [{'name': k['region'], 'count': k['count']} for k in ongoing_kpis],
    }


@reports_bp.route('/projects-dashboard')
@login_required
def projects_dashboard():
    data = _project_pivot_data()
    return render_template(
        'reports/projects_dashboard.html',
        now=date.today().strftime('%Y-%m-%d'),
        regions=data['regions'],
        statuses=data['statuses'],
        pivot=data['pivot'],
        region_counts=data['region_counts'],
        region_msar=data['region_msar'],
        count_row_totals=data['count_row_totals'],
        msar_row_totals=data['msar_row_totals'],
        total_projects=data['total_projects'],
        total_msar=data['total_msar'],
        ongoing_kpis=data['ongoing_kpis'],
        ongoing_count=data['ongoing_count'],
        ongoing_msar=data['ongoing_msar'],
        value_chart=json.dumps(data['value_chart'], ensure_ascii=False),
        count_chart=json.dumps(data['count_chart'], ensure_ascii=False),
        logo_alamro=_b64_logo('Alamro_Logo.png'),
        logo_nwc=_b64_logo('NWC_Logo.png'),
    )
@reports_bp.route('/finance')
@login_required
def finance():
    """Finance report: Complete PO, Invoices, and Variation Orders breakdown."""
    try:
        import openpyxl, shutil, os
        from pathlib import Path
        from flask import current_app

        # Data sources from config (supports both local paths and environment overrides)
        config = current_app.config
        PO_FILE = Path(config['EXCEL_SOURCES']['po_master'])
        PRO2 = Path(config['EXCEL_SOURCES']['invoices'])
        PO6_FILE = Path(config['EXCEL_SOURCES']['po6_detail'])
        VAR_FILE = Path(config['EXCEL_SOURCES']['variations'])

        ORIG_CONTRACT = 126_000_000
        EXT_CONTRACT = 305_111_979
        VAR_VALUE = EXT_CONTRACT - ORIG_CONTRACT

        def load_copy(path):
            if not path.exists():
                return None
            tmp = str(path) + ".tmp.xlsx"
            shutil.copy2(str(path), tmp)
            wb = openpyxl.load_workbook(tmp, data_only=True)
            os.remove(tmp)
            return wb

        def sv(v):
            if v is None: return ''
            return str(v).strip()

        def fmt(n):
            try:
                return f"{float(n):,.0f}"
            except:
                return '—'

        # ── PO 1-5 ────────────────────────────────────────────────────────────────
        po_list = []
        wb_po = load_copy(PO_FILE)
        if wb_po:
            ws_po = wb_po['Sheet1']
            for r in ws_po.iter_rows(min_row=2, max_row=6, values_only=True):
                if r and r[1] and 'PO' in str(r[1]):
                    po_list.append({
                        'name': sv(r[1]),
                        'amount': float(r[0] or 0)
                    })

        # ── Invoices (split: original vs variation) ────────────────────────────────
        orig_invoices = []
        var_invoices = []
        wb2 = load_copy(PRO2)
        if wb2:
            try:
                ws2 = wb2['إجمالي المستخلصات_']
            except:
                ws2 = wb2.active

            for r in ws2.iter_rows(min_row=5, max_row=61, values_only=True):
                if not r or not r[0]: continue
                label = sv(r[0])
                if 'مستخلص' not in label: continue

                po_no = r[1] if len(r) > 1 else None
                month = sv(r[2]) if len(r) > 2 else ''
                gross = float(r[3] or 0) if len(r) > 3 else 0
                ret10 = float(r[4] or 0) if len(r) > 4 and r[4] else 0
                vat = float(r[6] or 0) if len(r) > 6 and r[6] else 0
                total = float(r[7] or 0) if len(r) > 7 and r[7] else gross
                status = sv(r[10]) if len(r) > 10 else ''

                inv = {
                    'label': label,
                    'po': po_no,
                    'month': month,
                    'gross': gross,
                    'ret10': ret10,
                    'vat': vat,
                    'total': total,
                    'status': status
                }

                if po_no == 6:
                    var_invoices.append(inv)
                else:
                    orig_invoices.append(inv)

        orig_total = sum(i['gross'] for i in orig_invoices)
        var_total = sum(i['gross'] for i in var_invoices)
        var_remaining = VAR_VALUE - var_total

        # ── PO6 Job Details ───────────────────────────────────────────────────────
        po6_jobs = []
        wb6 = load_copy(PO6_FILE)
        if wb6:
            try:
                ws6 = wb6['ToTal_From PO_6_underway']
            except:
                ws6 = wb6.active

            for r in ws6.iter_rows(min_row=6, max_row=35, values_only=True):
                if r and r[0] is not None and isinstance(r[0], int):
                    po6_jobs.append({
                        'no': int(r[0]),
                        'desc': sv(r[1]) if len(r) > 1 else '',
                        'unit_price': float(r[3] or 0) if len(r) > 3 else 0,
                        'persons': float(r[4] or 0) if len(r) > 4 else 0,
                        'contract_months': float(r[5] or 0) if len(r) > 5 else 0,
                        'contract_qty': float(r[6] or 0) if len(r) > 6 else 0,
                        'contract_total': float(r[7] or 0) if len(r) > 7 else 0,
                        'cum_qty': float(r[10] or 0) if len(r) > 10 else 0,
                        'cum_total': float(r[11] or 0) if len(r) > 11 else 0,
                    })

        # ── Variation Budget per Job ──────────────────────────────────────────────
        var_budget = {}
        wb_var = load_copy(VAR_FILE)
        if wb_var:
            try:
                ws_var = wb_var['1']
            except:
                ws_var = wb_var.active

            for r in ws_var.iter_rows(min_row=3, max_row=32, values_only=True):
                if r and r[0] and isinstance(r[0], int):
                    orig_v = float(r[7] or 0) if len(r) > 7 else 0
                    amend_v = float(r[11] or 0) if len(r) > 11 else 0
                    var_budget[int(r[0])] = max(amend_v - orig_v, 0)

        # Enrich PO6 jobs with variation budget
        for job in po6_jobs:
            job['var_budget'] = var_budget.get(job['no'], 0)
            job['var_remaining'] = max(job['var_budget'] - job['cum_total'], 0)
            job['pct_complete'] = (job['cum_total'] / job['var_budget'] * 100) if job['var_budget'] > 0 else 0

        # ── Summary Data ───────────────────────────────────────────────────────────
        po_total_allocated = sum(p['amount'] for p in po_list)
        po_total_disbursed = orig_total

        return render_template(
            'reports/finance.html',
            po_list=po_list,
            orig_invoices=orig_invoices,
            var_invoices=var_invoices,
            po6_jobs=po6_jobs,
            kpis_orig={
                'allocated': po_total_allocated,
                'disbursed': po_total_disbursed,
                'remaining': po_total_allocated - po_total_disbursed,
            },
            kpis_var={
                'allocated': VAR_VALUE,
                'disbursed': var_total,
                'remaining': var_remaining,
            },
            logo_nwc=_b64_logo('NWC_Logo.png'),
            logo_alamro=_b64_logo('Alamro_Logo.png'),
        )

    except Exception as e:
        import traceback
        return render_template('reports/finance.html',
                              error=str(e),
                              traceback=traceback.format_exc(),
                              logo_nwc=_b64_logo('NWC_Logo.png'),
                              logo_alamro=_b64_logo('Alamro_Logo.png'))


# ──────────────────────────────────────────────────────────────────────────────
# Org Chart Routes (Tasks 8-10)
# ──────────────────────────────────────────────────────────────────────────────

ORG_CHART_REGION_MAP = {
    'asir': {'ar': 'عسير', 'file': '09_OrgChart_Asir.html'},
    'jizan': {'ar': 'جازان', 'file': '10_OrgChart_Jizan.html'},
    'baha': {'ar': 'الباحة', 'file': '11_OrgChart_Baha.html'},
    'najran': {'ar': 'نجران', 'file': '12_OrgChart_Najran.html'},
}


@reports_bp.route('/org-chart')
@login_required
def org_chart_landing():
    """Landing page: 4 region cards."""
    regions = [
        {'code': 'asir', 'name': 'عسير'},
        {'code': 'jizan', 'name': 'جازان'},
        {'code': 'baha', 'name': 'الباحة'},
        {'code': 'najran', 'name': 'نجران'},
    ]
    return render_template('reports/org_chart_landing.html', regions=regions)


@reports_bp.route('/org-chart/<region>')
@login_required
def org_chart_view(region):
    """View org chart for a region."""
    if region not in ORG_CHART_REGION_MAP:
        return render_template('error.html',
                              message='منطقة غير صحيحة'), 404

    info = ORG_CHART_REGION_MAP[region]
    org_chart_file = Path(__file__).parent.parent.parent / 'static' / 'org_charts' / info['file']

    if not org_chart_file.exists():
        return render_template('error.html',
                              message='ملف الهيكل التنظيمي غير متوفر'), 404

    html_content = org_chart_file.read_text(encoding='utf-8')

    return render_template('reports/org_chart_view.html',
                          region=region,
                          region_name=info['ar'],
                          org_chart_html=html_content)


@reports_bp.route('/org-chart/<region>/pdf')
@login_required
def org_chart_pdf(region):
    """Export org chart as PDF."""
    if region not in ORG_CHART_REGION_MAP:
        return {'error': 'منطقة غير صحيحة'}, 404

    try:
        from playwright.sync_api import sync_playwright
        import tempfile
        import os

        info = ORG_CHART_REGION_MAP[region]
        org_chart_file = Path(__file__).parent.parent.parent / 'static' / 'org_charts' / info['file']

        if not org_chart_file.exists():
            return {'error': 'ملف الهيكل التنظيمي غير متوفر'}, 404

        html_content = org_chart_file.read_text(encoding='utf-8')

        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as tmp_html:
            tmp_html.write(html_content)
            tmp_html_path = tmp_html.name

        try:
            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page()
                page.goto(f"file://{tmp_html_path}")

                pdf_bytes = page.pdf(
                    format='A3',
                    landscape=True,
                    margin={'top': '10mm', 'bottom': '10mm', 'left': '14mm', 'right': '14mm'}
                )

                browser.close()

            os.unlink(tmp_html_path)

            from flask import send_file
            import io

            now = date.today().strftime('%Y-%m-%d')
            filename = f"OrgChart_{info['ar']}_{now}.pdf"

            return send_file(
                io.BytesIO(pdf_bytes),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )

        except Exception as e:
            if os.path.exists(tmp_html_path):
                os.unlink(tmp_html_path)
            raise

    except ImportError:
        return {'error': 'Playwright غير متوفر. لا يمكن تصدير PDF'}, 503
    except Exception as e:
        return {'error': f'خطأ في إنشاء PDF: {str(e)}'}, 500
