"""
DB → Excel export. Writes source/employees data source.xlsx (sheet 'data')
in the exact original column layout so all existing gen_*.py scripts work unchanged.

Excel column map (0-based, sheet 'data') — mirrors import_excel.py exactly:
  0  = المؤهل الدراسي  (qualification)
  1  = سنة التخرج      (grad_year)
  2  = الايميل         (email)
  3  = رقم الجوال      (phone)
  4  = تاريخ المباشرة  (hire_date)
  5  = رقم الهوية      (id_number)
  6  = عدد سنوات الخبرة (experience_years)
  7  = استيفاء المتطلبات (not stored — left blank)
  8  = المهنة          (profession)
  9  = تاريخ انتهاء العقد (contract_end_date)
  10 = تاريخ الاستقالة  (resignation_date)
  11 = الموقف الحالي   (status)
  12 = السبب           (status_reason)
  13 = تقييم المدير المباشر (not stored — left blank)
  14 = الجنسية         (nationality name_ar)
  15 = الكفالة         (kafala)
  16 = المنطقة         (region)
  17 = CBU             (cbu)
  18 = إشراف - دعم فني (category / supervision_type)
  19 = المكتب          (office name)
  20 = الاسم           (full_name)
  21 = الوظيفة         (job_code title)
  22 = الكود2          (serial_no)
  23 = سعر الوحدة      (unit_price)
  24 = المدير المباشر  (direct_manager)
  25 = كشف حضور       (attendance_group name)
  26 = رقم_مسلسل      (row number — written as sequential int)
  27 = RE              (re_code)
  28 = م1              (not stored — left blank)
  29 = كشف حضور يوليو (not stored — left blank)

Run from hr_webapp/: python scripts/export_emp_excel.py
"""
import sys, os, shutil
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

HR_ROOT = Path(__file__).parent.parent.parent.parent.parent
EXCEL   = HR_ROOT / 'source' / 'employees data source.xlsx'
STATUS_ACTIVE = 'على قوة العمل'


def export_to_excel(app=None, excel_path=None):
    """
    Export active employees from DB to Excel.

    Pass app to run standalone; call from within an active app context otherwise.
    Pass excel_path to override the default EXCEL path (used in tests).
    Returns count of employees written.
    """
    target = Path(excel_path) if excel_path else EXCEL
    _standalone = app is not None
    if _standalone:
        ctx = app.app_context()
        ctx.push()

    if not target.exists():
        raise FileNotFoundError(
            f'Excel source not found: {target}\n'
            'Run the export only after the source file exists at that path.'
        )

    try:
        import openpyxl
        from app.models import Employee, EmployeeStatus
        from app import db

        employees = (
            Employee.query
            .join(Employee.current_status)
            .filter(EmployeeStatus.name_ar == STATUS_ACTIVE)
            .order_by(Employee.region, Employee.full_name)
            .all()
        )

        # Load original workbook to preserve other sheets and header row
        tmp = str(target) + '.export_tmp.xlsx'
        shutil.copy2(str(target), tmp)
        import openpyxl as _opxl
        wb = _opxl.load_workbook(tmp)
        os.remove(tmp)
        ws = wb['data']
        max_col = ws.max_column

        # Ensure we have at least 30 columns (indices 0-29)
        needed_cols = 30
        if max_col < needed_cols:
            max_col = needed_cols

        # Clear data rows (keep header row 1)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Write employees — reverse mapping of import_excel.py column assignments
        for r_idx, emp in enumerate(employees, start=2):
            row_data = [None] * max_col

            # --- Columns that were NOT stored (left as None / blank) ---
            # col 7  = استيفاء المتطلبات  — not imported, leave blank
            # col 13 = تقييم المدير المباشر — not imported, leave blank
            # col 28 = م1                  — not imported, leave blank
            # col 29 = كشف حضور يوليو     — monthly snapshot, leave blank

            # --- Core fields (always written) ---
            row_data[11] = STATUS_ACTIVE
            row_data[20] = emp.full_name
            row_data[3]  = emp.phone or ''
            row_data[14] = emp.nationality.name_ar if emp.nationality else ''
            row_data[15] = emp.kafala or ''
            row_data[16] = emp.region or ''
            row_data[17] = emp.cbu or ''
            row_data[18] = emp.category or emp.supervision_type or ''
            row_data[21] = emp.job_code.title if emp.job_code else ''
            row_data[23] = float(emp.unit_price) if emp.unit_price else ''

            # --- Additional fields discovered in Task 9 import ---
            row_data[0]  = emp.qualification or ''
            row_data[1]  = emp.grad_year if emp.grad_year is not None else ''
            row_data[2]  = emp.email or ''
            row_data[4]  = emp.hire_date.isoformat() if emp.hire_date else ''
            row_data[5]  = emp.id_number or ''
            row_data[6]  = emp.experience_years if emp.experience_years is not None else ''
            row_data[8]  = emp.profession or ''
            row_data[9]  = emp.contract_end_date.isoformat() if emp.contract_end_date else ''
            row_data[10] = emp.resignation_date.isoformat() if emp.resignation_date else ''
            row_data[12] = emp.status_reason or ''
            row_data[19] = emp.office.name if emp.office else ''
            row_data[22] = emp.serial_no or ''
            row_data[24] = emp.direct_manager or ''
            row_data[25] = emp.attendance_group.name if emp.attendance_group else ''
            row_data[26] = r_idx - 1   # sequential row number (1-based relative to data)
            row_data[27] = emp.re_code or ''

            for c_idx, val in enumerate(row_data, start=1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)

        # Backup then save
        backup = str(target) + f'.bak_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        shutil.copy2(str(target), backup)
        wb.save(str(target))

        count = len(employees)
        print(f'Export complete. {count} employees written to {target.name}')
        print(f'Backup saved: {Path(backup).name}')
        return count

    finally:
        if _standalone:
            ctx.pop()


if __name__ == '__main__':
    from app import create_app
    export_to_excel(app=create_app())
