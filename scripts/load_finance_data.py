#!/usr/bin/env python3
"""
load_finance_data.py — Load Finance Report data from Excel into database

Usage:
    python scripts/load_finance_data.py

This script:
1. Reads Excel source files from config.EXCEL_SOURCES
2. Parses PO, Invoice, and PO6 job data
3. Loads into FinancePO, FinanceInvoice, FinancePO6Job tables
4. Enables fast report queries from database instead of Excel files
"""

import sys
import os
from pathlib import Path

# Add parent to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app import create_app, db
from app.models import FinancePO, FinanceInvoice, FinancePO6Job
import openpyxl
import shutil


def load_copy(path):
    """Load Excel file safely (copy to temp, load, delete temp)."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    tmp = str(path) + ".tmp.xlsx"
    shutil.copy2(str(path), tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
        return wb
    finally:
        if Path(tmp).exists():
            os.remove(tmp)


def sv(v):
    """Safe value: convert None to empty string, skip Excel errors."""
    if v is None:
        return ''
    s = str(v).strip()
    if s in {'#REF!', '=#REF!', '#VALUE!', '#N/A', '#NAME?', '#DIV/0!', '#NULL!', '#NUM!'}:
        return ''
    return s


def load_po_data(excel_path):
    """Load PO 1-5 from Excel."""
    print(f"  Loading PO data from {Path(excel_path).name}...")
    wb = load_copy(excel_path)
    ws = wb['Sheet1']
    count = 0

    for r in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        if r and len(r) > 1 and r[1] and 'PO' in str(r[1]):
            po = FinancePO(
                po_number=sv(r[1]),
                allocated_amount=float(r[0] or 0),
                status='completed'
            )
            db.session.merge(po)
            count += 1

    db.session.commit()
    print(f"  ✓ Loaded {count} purchase orders")


def load_invoice_data(excel_path):
    """Load invoices from Excel."""
    print(f"  Loading invoice data from {Path(excel_path).name}...")
    wb = load_copy(excel_path)
    ws = wb['إجمالي المستخلصات_']
    count = 0

    for r in ws.iter_rows(min_row=5, max_row=61, values_only=True):
        if not r or not r[0]:
            continue
        label = sv(r[0])
        if 'مستخلص' not in label:
            continue

        po_no = r[1] if len(r) > 1 else None
        month = sv(r[2]) if len(r) > 2 else ''
        gross = float(r[3] or 0) if len(r) > 3 else 0
        ret10 = float(r[4] or 0) if len(r) > 4 and r[4] else 0
        vat = float(r[6] or 0) if len(r) > 6 and r[6] else 0
        total = float(r[7] or 0) if len(r) > 7 and r[7] else gross
        status = sv(r[10]) if len(r) > 10 else ''

        inv = FinanceInvoice(
            invoice_label=label,
            po_number=po_no,
            month=month,
            gross_amount=gross,
            retention_10=ret10,
            vat_amount=vat,
            net_amount=total,
            status=status
        )
        db.session.merge(inv)
        count += 1

    db.session.commit()
    print(f"  ✓ Loaded {count} invoices")


def load_po6_jobs(excel_path, variations_path):
    """Load PO 6 job details and variation budgets."""
    print(f"  Loading PO 6 jobs from {Path(excel_path).name}...")

    # Load variation budgets
    var_budget = {}
    wb_var = load_copy(variations_path)
    ws_var = wb_var['1']
    for r in ws_var.iter_rows(min_row=3, max_row=32, values_only=True):
        if r and r[0] and isinstance(r[0], int):
            orig_v = float(r[7] or 0) if len(r) > 7 else 0
            amend_v = float(r[11] or 0) if len(r) > 11 else 0
            var_budget[int(r[0])] = max(amend_v - orig_v, 0)

    # Load jobs
    wb = load_copy(excel_path)
    ws = wb['ToTal_From PO_6_underway']
    count = 0

    for r in ws.iter_rows(min_row=6, max_row=35, values_only=True):
        if r and r[0] is not None and isinstance(r[0], int):
            job = FinancePO6Job(
                job_no=int(r[0]),
                description=sv(r[1]) if len(r) > 1 else '',
                unit_price=float(r[3] or 0) if len(r) > 3 else 0,
                persons=float(r[4] or 0) if len(r) > 4 else 0,
                contract_months=float(r[5] or 0) if len(r) > 5 else 0,
                contract_qty=float(r[6] or 0) if len(r) > 6 else 0,
                contract_total=float(r[7] or 0) if len(r) > 7 else 0,
                cumulative_qty=float(r[10] or 0) if len(r) > 10 else 0,
                cumulative_total=float(r[11] or 0) if len(r) > 11 else 0,
                variation_budget=var_budget.get(int(r[0]), 0)
            )
            db.session.merge(job)
            count += 1

    db.session.commit()
    print(f"  ✓ Loaded {count} PO 6 jobs with variation budgets")


def main():
    """Load all finance data from Excel sources."""
    app = create_app('development')

    with app.app_context():
        print("\n📊 Loading Finance Report data from Excel sources...\n")

        try:
            sources = app.config['EXCEL_SOURCES']

            # Clear existing data
            print("  Clearing existing data...")
            db.session.query(FinancePO).delete()
            db.session.query(FinanceInvoice).delete()
            db.session.query(FinancePO6Job).delete()
            db.session.commit()
            print("  ✓ Cleared\n")

            # Load all data
            load_po_data(sources['po_master'])
            load_invoice_data(sources['invoices'])
            load_po6_jobs(sources['po6_detail'], sources['variations'])

            print("\n✅ Finance data loaded successfully!\n")
            print("Reports now read from database (fast, cached) instead of Excel.")
            print("To refresh data: python scripts/load_finance_data.py\n")

        except Exception as e:
            print(f"\n❌ Error loading data: {e}\n")
            raise


if __name__ == '__main__':
    main()
