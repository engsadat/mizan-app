"""
Import projects from Excel source.
Usage: python scripts/import_projects.py [path/to/source.xlsx]
Default: C:/Users/engsa/OneDrive/Desktop/AI/HR/source/project_2026_database_ver1_updated.xlsx
"""
import sys, shutil, os
from pathlib import Path
from datetime import datetime
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
from app import create_app, db
from app.models import Project

DEFAULT_SRC = Path("C:/Users/engsa/OneDrive/Desktop/AI/HR/source/project_2026_database_ver1_updated.xlsx")
src_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC

print(f"Source: {src_path}")
if not src_path.exists():
    print("ERROR: file not found"); sys.exit(1)

# Safe load with copy
tmp = Path(str(src_path) + ".tmp.xlsx")
shutil.copy2(src_path, tmp)
wb = openpyxl.load_workbook(tmp, data_only=True)
os.remove(tmp)

ws = wb['pro']
print(f"Sheet 'pro': {ws.max_row} rows, {ws.max_column} cols\n")

# Column mapping (0-indexed from user's list)
COL = {
    'x': 2, 'y': 3,
    'network_approved': 4,  # col E
    'network_implemented': 5,  # col F
    'operational_contract_no': 7,  # col H
    'project_number': 8,  # col I
    'po_no': 9,  # col J
    'included': 10,  # col K
    'name': 11,  # col L (إسم المشروع)
    'value': 28,  # col AC (القيمة)
    'description': 12,  # col M (الوصف)
    'funding_source': 13,  # col N (مصدر التمويل)
    'city': 14,  # col O (المدينة)
    'region': 15,  # col P (المنطقة)
    'contractor_name': 17,  # col R (اسم المقاول)
    'is_shared': 18,  # col S (مشترك)
    're_code': 19,  # col T (م مقيم)
    'project_status': 20,  # col U (وضع المشروع)
    'project_state': 21,  # col V (حالة المشروع)
    'start_date': 22,  # col W (تاريخ البدء)
    'end_date': 23,  # col X (تاريخ الانتهاء)
    'contractor': 26,  # col AA (Contractor)
    'classification': 27,  # col AB (تصنيف المشروع)
    're_asir': 29,  # col AD (اسم المقيم- عسير)
    're_jazan': 30,  # col AE (اسم المقيم- جازان)
    're_baha': 31,  # col AF (اسم المقيم-الباحة)
    're_najran': 32,  # col AG (اسم المقيم- نجران)
    'planned_completion': 33,  # col AH (نسبة الإنجاز المخططة)
    'actual_completion': 34,  # col AI (نسبة الإنجاز الفعلية)
    'variance': 35,  # col AJ (الفارق)
    'treatment_stations': 36,  # col AK (عدد محطة معالجة)
    'pumping_stations': 37,  # col AL (عدد محطة رفع/ ضخ)
}

def to_str(v):
    if v is None: return ''
    return str(v).strip()

def to_bool(v):
    if v is None: return False
    s = str(v).lower().strip()
    return s in ('yes', 'نعم', 'true', '1')

def to_float(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    try:
        if isinstance(v, str):
            # Try multiple date formats
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                try: return datetime.strptime(v, fmt).date()
                except: pass
        return None
    except: return None

app = create_app()
with app.app_context():
    db.create_all()

    # Clear existing
    Project.query.delete()
    db.session.commit()
    print("Cleared existing projects\n")

    imported = 0
    errors = []

    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]

        try:
            # Get name (required)
            name = to_str(row[COL['name']].value)
            if not name:
                continue

            # Build project
            proj = Project(
                x = to_float(row[COL['x']].value),
                y = to_float(row[COL['y']].value),
                network_approved = to_float(row[COL['network_approved']].value),
                network_implemented = to_float(row[COL['network_implemented']].value),
                operational_contract_no = to_str(row[COL['operational_contract_no']].value),
                project_number = to_str(row[COL['project_number']].value),
                po_no = to_str(row[COL['po_no']].value),
                included = to_bool(row[COL['included']].value),
                name = name,
                value = to_float(row[COL['value']].value),
                description = to_str(row[COL['description']].value),
                funding_source = to_str(row[COL['funding_source']].value),
                city = to_str(row[COL['city']].value),
                region = to_str(row[COL['region']].value),
                contractor_name = to_str(row[COL['contractor_name']].value),
                is_shared = to_bool(row[COL['is_shared']].value),
                re_code = to_str(row[COL['re_code']].value),
                project_status = to_str(row[COL['project_status']].value),
                project_state = to_str(row[COL['project_state']].value),
                start_date = to_date(row[COL['start_date']].value),
                end_date = to_date(row[COL['end_date']].value),
                contractor = to_str(row[COL['contractor']].value),
                classification = to_str(row[COL['classification']].value),
                re_asir = to_str(row[COL['re_asir']].value),
                re_jazan = to_str(row[COL['re_jazan']].value),
                re_baha = to_str(row[COL['re_baha']].value),
                re_najran = to_str(row[COL['re_najran']].value),
                planned_completion = to_float(row[COL['planned_completion']].value),
                actual_completion = to_float(row[COL['actual_completion']].value),
                variance = to_float(row[COL['variance']].value),
                treatment_stations = int(to_float(row[COL['treatment_stations']].value)),
                pumping_stations = int(to_float(row[COL['pumping_stations']].value)),
            )
            db.session.add(proj)
            imported += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    db.session.commit()
    print(f"Imported: {imported} projects")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for err in errors[:10]:
            print(f"  {err}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more")
