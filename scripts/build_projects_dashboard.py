#!/usr/bin/env python3
"""
Standalone projects-dashboard PREVIEW.

Not part of the Flask app. Edit here, open reports_out/projects_dashboard.html.
When you approve: copy the layout into app/templates/reports/.

  python scripts/build_projects_dashboard.py --demo
  python scripts/build_projects_dashboard.py path\\to\\project_2026.xlsx

Reads Excel sheet ``pro``, Included = Yes only (same rule as على قوة العمل).
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import matplotlib

matplotlib.use('Agg')
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import arabic_reshaper
from bidi.algorithm import get_display

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / 'reports_out'
DEFAULT_XLSX = Path(
    os.environ.get(
        'PROJECTS_EXCEL',
        r'C:\Users\engsa\OneDrive\Desktop\AI\HR\source\project_2026_database_ver1_updated.xlsx',
    )
)

# Same 0-based columns as scripts/import_projects.py
COL = {
    'included': 10,
    'name': 11,
    'value': 28,
    'region': 15,
    'project_state': 21,
    'planned': 33,
    'actual': 34,
    'classification': 27,
}

REGIONS = ['عسير', 'جازان', 'الباحة', 'نجران']
NAVY, TEAL, AMBER, RED, SLATE = '#0a1f3d', '#0071b9', '#d97706', '#b91c1c', '#64748b'
STATUS_COLORS = {
    'جارٍ': TEAL,
    'مكتمل': '#059669',
    'متعثر': RED,
    'لم يبدأ': '#7c3aed',
    'غير محدد': SLATE,
}
AR_FONTS = [
    '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf',
    r'C:\Windows\Fonts\segoeui.ttf',
    r'C:\Windows\Fonts\tahoma.ttf',
    r'C:\Windows\Fonts\arial.ttf',
]


def font_path() -> str | None:
    for path in AR_FONTS:
        if Path(path).exists():
            return path
    return None


def fp() -> fm.FontProperties:
    p = font_path()
    return fm.FontProperties(fname=p) if p else fm.FontProperties()


def ar(text) -> str:
    """Arabic for matplotlib.

    Linux + Noto already shapes glyphs (HarfBuzz). python-bidi then
    double-reverses the string. Windows fonts still need reshape + bidi.
    """
    s = '' if text is None else str(text)
    if not s:
        return s
    if sys.platform.startswith('win'):
        return get_display(arabic_reshaper.reshape(s))
    return s


def to_str(v) -> str:
    return '' if v is None else str(v).strip()


def to_bool(v) -> bool:
    return str(v).lower().strip() in ('yes', 'نعم', 'true', '1')


def to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', '').strip() or 0)
    except ValueError:
        return 0.0


def demo_rows():
    """Layout only — not production numbers."""
    return [
        {'name': 'شبكات مياه أبها', 'region': 'عسير', 'state': 'جارٍ',
         'value': 420e6, 'planned': 55, 'actual': 38, 'classification': 'شبكات مياه'},
        {'name': 'صرف صحي خميس مشيط', 'region': 'عسير', 'state': 'مكتمل',
         'value': 180e6, 'planned': 100, 'actual': 100, 'classification': 'شبكات صرف'},
        {'name': 'محطة معالجة أبها', 'region': 'عسير', 'state': 'جارٍ',
         'value': 88e6, 'planned': 45, 'actual': 48, 'classification': 'محطات معالجة'},
        {'name': 'شبكات جازان', 'region': 'جازان', 'state': 'جارٍ',
         'value': 310e6, 'planned': 70, 'actual': 41, 'classification': 'شبكات مياه'},
        {'name': 'رفع صبيا', 'region': 'جازان', 'state': 'متعثر',
         'value': 95e6, 'planned': 80, 'actual': 22, 'classification': 'محطات رفع'},
        {'name': 'شبكات الباحة', 'region': 'الباحة', 'state': 'جارٍ',
         'value': 140e6, 'planned': 40, 'actual': 36, 'classification': 'شبكات مياه'},
        {'name': 'صرف نجران', 'region': 'نجران', 'state': 'جارٍ',
         'value': 210e6, 'planned': 60, 'actual': 44, 'classification': 'شبكات صرف'},
        {'name': 'خزان نجران', 'region': 'نجران', 'state': 'لم يبدأ',
         'value': 75e6, 'planned': 10, 'actual': 0, 'classification': 'خزانات'},
    ]


def load_excel(path: Path):
    import openpyxl
    tmp = Path(str(path) + '.tmp.xlsx')
    shutil.copy2(path, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
        ws = wb['pro'] if 'pro' in wb.sheetnames else wb.active
        rows = []
        for i in range(2, ws.max_row + 1):
            r = ws[i]
            if not to_bool(r[COL['included']].value):
                continue
            name = to_str(r[COL['name']].value)
            if not name:
                continue
            rows.append({
                'name': name,
                'region': to_str(r[COL['region']].value) or 'غير محدد',
                'state': to_str(r[COL['project_state']].value) or 'غير محدد',
                'value': to_float(r[COL['value']].value),
                'planned': to_float(r[COL['planned']].value),
                'actual': to_float(r[COL['actual']].value),
                'classification': to_str(r[COL['classification']].value) or 'غير محدد',
            })
        return rows
    finally:
        if tmp.exists():
            tmp.unlink()


def summarize(rows):
    by_region_val = defaultdict(float)
    by_region_plan = defaultdict(list)
    by_region_act = defaultdict(list)
    by_state = defaultdict(int)
    by_class = defaultdict(int)
    pivot = defaultdict(lambda: defaultdict(lambda: {'n': 0, 'value': 0.0}))
    behind = 0
    for p in rows:
        by_region_val[p['region']] += p['value']
        by_region_plan[p['region']].append(p['planned'])
        by_region_act[p['region']].append(p['actual'])
        by_state[p['state']] += 1
        by_class[p['classification']] += 1
        pivot[p['state']][p['region']]['n'] += 1
        pivot[p['state']][p['region']]['value'] += p['value']
        if p['planned'] > 0 and p['actual'] < p['planned']:
            behind += 1
    n = len(rows)
    total_value = sum(p['value'] for p in rows)
    avg_act = (sum(p['actual'] for p in rows) / n) if n else 0
    regions = [r for r in REGIONS if r in by_region_val] or sorted(by_region_val)
    states = list(by_state.keys())
    return {
        'n': n,
        'total_value': total_value,
        'avg_act': avg_act,
        'behind': behind,
        'regions': regions,
        'region_values': [by_region_val[r] for r in regions],
        'region_plan': [
            sum(by_region_plan[r]) / len(by_region_plan[r]) if by_region_plan[r] else 0
            for r in regions
        ],
        'region_act': [
            sum(by_region_act[r]) / len(by_region_act[r]) if by_region_act[r] else 0
            for r in regions
        ],
        'states': states,
        'state_counts': [by_state[s] for s in states],
        'class_rows': sorted(by_class.items(), key=lambda x: -x[1])[:10],
        'pivot': pivot,
        'region_totals': dict(by_region_val),
    }


def _setup_fonts():
    plt.rcParams['font.family'] = 'DejaVu Sans'
    plt.rcParams['axes.unicode_minus'] = False


def style_ax(ax, title: str):
    f = fp()
    ax.set_title(ar(title), fontproperties=f, fontsize=13, color=NAVY, pad=12)
    ax.tick_params(colors=NAVY)
    for lab in ax.get_xticklabels() + ax.get_yticklabels():
        lab.set_fontproperties(f)
        lab.set_fontsize(10)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.set_facecolor('white')
    ax.yaxis.grid(True, color='#e2e8f0', zorder=0)
    ax.set_axisbelow(True)


def save_charts(s, demo: bool):
    _setup_fonts()
    OUT.mkdir(parents=True, exist_ok=True)
    f = fp()
    note = ar('معاينة تصميم ارقام تجريبية') if demo else ar('المشاريع المدرجة من ورقة pro')
    paths = []

    def footer(fig):
        fig.text(0.99, 0.015, note, ha='right', fontproperties=f, fontsize=8, color=SLATE)

    # 1) value by region
    fig, ax = plt.subplots(figsize=(8.4, 4.5), dpi=150)
    fig.patch.set_facecolor('white')
    vals_m = [v / 1e6 for v in s['region_values']]
    bars = ax.bar([ar(r) for r in s['regions']], vals_m, color=NAVY, width=0.58, zorder=3)
    for b, v in zip(bars, vals_m):
        ax.text(
            b.get_x() + b.get_width() / 2, b.get_height(),
            f'{v:.0f}', ha='center', va='bottom', fontsize=10,
            color=NAVY, fontweight='bold',
        )
    ax.set_ylabel(ar('مليون ريال'), fontproperties=f, color=NAVY)
    style_ax(ax, 'القيمة حسب المنطقة')
    footer(fig)
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    p1 = OUT / 'chart_region_value.png'
    fig.savefig(p1, bbox_inches='tight')
    plt.close(fig)
    paths.append(p1)

    # 2) count by status
    fig, ax = plt.subplots(figsize=(8.4, 4.5), dpi=150)
    fig.patch.set_facecolor('white')
    colors = [STATUS_COLORS.get(st, TEAL) for st in s['states']]
    bars = ax.bar(
        [ar(x) for x in s['states']], s['state_counts'],
        color=colors, width=0.58, zorder=3,
    )
    for b, v in zip(bars, s['state_counts']):
        ax.text(
            b.get_x() + b.get_width() / 2, b.get_height(),
            str(v), ha='center', va='bottom', fontsize=11,
            color=NAVY, fontweight='bold',
        )
    ax.set_ylabel(ar('عدد المشاريع'), fontproperties=f, color=NAVY)
    style_ax(ax, 'عدد المشاريع حسب الحالة')
    footer(fig)
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    p2 = OUT / 'chart_status_count.png'
    fig.savefig(p2, bbox_inches='tight')
    plt.close(fig)
    paths.append(p2)

    # 3) planned vs actual by region
    fig, ax = plt.subplots(figsize=(8.8, 4.6), dpi=150)
    fig.patch.set_facecolor('white')
    x = list(range(len(s['regions'])))
    w = 0.36
    ax.bar([i - w / 2 for i in x], s['region_plan'], width=w, color=TEAL, label=ar('مخطط'), zorder=3)
    ax.bar([i + w / 2 for i in x], s['region_act'], width=w, color=AMBER, label=ar('فعلي'), zorder=3)
    ax.set_xticks(x)
    ax.set_xticklabels([ar(r) for r in s['regions']])
    ax.set_ylabel(ar('نسبة الإنجاز'), fontproperties=f, color=NAVY)
    ax.set_ylim(0, 115)
    ax.legend(prop=f, frameon=False, loc='upper right')
    style_ax(ax, 'الإنجاز المخطط مقابل الفعلي حسب المنطقة')
    footer(fig)
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    p3 = OUT / 'chart_progress.png'
    fig.savefig(p3, bbox_inches='tight')
    plt.close(fig)
    paths.append(p3)

    return paths


def write_html(s, demo: bool):
    banner = 'معاينة تصميم — أرقام تجريبية' if demo else 'المشاريع المدرجة من ورقة pro'
    region_chart = [
        {'name': r, 'value': v}
        for r, v in zip(s['regions'], s['region_values'])
    ]
    status_chart = [
        {'name': n, 'count': c}
        for n, c in zip(s['states'], s['state_counts'])
    ]
    progress_chart = [
        {'name': r, 'planned': p, 'actual': a}
        for r, p, a in zip(s['regions'], s['region_plan'], s['region_act'])
    ]

    pivot_rows = []
    for status in s['states']:
        cells = []
        row_total = 0.0
        for region in s['regions']:
            val = s['pivot'][status][region]['value']
            cells.append(val)
            row_total += val
        pivot_rows.append({'status': status, 'cells': cells, 'total': row_total})

    def money(v):
        return f'{v:,.0f}'

    table_body = []
    for row in pivot_rows:
        tds = ''.join(f'<td class="num">{money(c)}</td>' for c in row['cells'])
        table_body.append(
            f'<tr><td>{row["status"]}</td>{tds}'
            f'<td class="num fw">{money(row["total"])}</td></tr>'
        )
    totals = ''.join(
        f'<td class="num">{money(s["region_totals"].get(r, 0))}</td>'
        for r in s['regions']
    )
    class_rows = ''.join(
        f'<tr><td>{name}</td><td class="num">{n}</td></tr>'
        for name, n in s['class_rows']
    )

    html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<title>ميزان — لوحة المشاريع (معاينة)</title>
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
<style>
body {{ font-family: Cairo, sans-serif; background:#f0f4f8; margin:0; color:{NAVY}; }}
.nav {{ background:{NAVY}; color:#fff; padding:12px 20px; font-weight:700; }}
.wrap {{ max-width:1100px; margin:0 auto; padding:20px; }}
.note {{ color:{SLATE}; font-size:.85rem; margin:8px 0 16px; }}
.kpis {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-bottom:16px; }}
.kpi {{ background:#fff; border-radius:12px; padding:16px; box-shadow:0 2px 8px rgba(0,0,0,.07); }}
.kpi b {{ display:block; font-size:1.55rem; }}
.kpi span {{ color:{SLATE}; font-size:.8rem; }}
.charts {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; }}
.card {{ background:#fff; border-radius:12px; padding:16px; box-shadow:0 2px 8px rgba(0,0,0,.07); }}
.card h2 {{ font-size:.9rem; margin:0 0 10px; border-right:4px solid {TEAL}; padding-right:8px; }}
.chart-box {{ height:240px; }}
.wide {{ margin-top:12px; }}
table {{ width:100%; border-collapse:collapse; font-size:.82rem; }}
th {{ background:{NAVY}; color:#fff; padding:8px; text-align:center; }}
td {{ border:1px solid #e2e8f0; padding:7px 8px; }}
td.num {{ text-align:center; font-family:Arial,sans-serif; }}
.fw {{ font-weight:700; }}
.foot {{ color:{SLATE}; font-size:.75rem; margin-top:16px; }}
@media (max-width:800px) {{
  .kpis, .charts {{ grid-template-columns:1fr 1fr; }}
}}
</style>
</head>
<body>
<div class="nav">ميزان — لوحة مشاريع القطاع الجنوبي</div>
<div class="wrap">
  <p class="note">{banner} · {s['n']} مشروع · القيمة {s['total_value']:,.0f} ر.س</p>
  <div class="kpis">
    <div class="kpi"><b style="color:#059669">{s['n']}</b><span>مشاريع مدرجة</span></div>
    <div class="kpi"><b>{s['total_value']/1e6:.0f} م</b><span>إجمالي القيمة</span></div>
    <div class="kpi"><b style="color:{TEAL}">{s['avg_act']:.0f}%</b><span>متوسط الإنجاز الفعلي</span></div>
    <div class="kpi"><b style="color:{AMBER}">{s['behind']}</b><span>متأخر عن المخطط</span></div>
  </div>
  <div class="charts">
    <div class="card"><h2>القيمة حسب المنطقة</h2><div class="chart-box"><canvas id="c1"></canvas></div></div>
    <div class="card"><h2>عدد المشاريع حسب الحالة</h2><div class="chart-box"><canvas id="c2"></canvas></div></div>
  </div>
  <div class="card wide"><h2>الإنجاز المخطط مقابل الفعلي حسب المنطقة</h2>
    <div class="chart-box"><canvas id="c3"></canvas></div>
  </div>
  <div class="card wide">
    <h2>القيمة حسب الحالة والمنطقة (ر.س)</h2>
    <table>
      <thead><tr><th style="text-align:right">حالة المشروع</th>
      {''.join(f'<th>{r}</th>' for r in s['regions'])}
      <th>الإجمالي</th></tr></thead>
      <tbody>
        {''.join(table_body)}
        <tr class="fw" style="background:#f1f5f9"><td>الإجمالي</td>{totals}
        <td class="num">{money(s['total_value'])}</td></tr>
      </tbody>
    </table>
  </div>
  <div class="card wide">
    <h2>التصنيف</h2>
    <table>
      <thead><tr><th style="text-align:right">التصنيف</th><th>العدد</th></tr></thead>
      <tbody>{class_rows}</tbody>
    </table>
  </div>
  <p class="foot">ملف مستقل للمعاينة فقط. بعد الموافقة تُنقل اللوحة إلى تطبيق ميزان ثم git.</p>
</div>
<script>
Chart.register(ChartDataLabels);
const NAVY = '{NAVY}', TEAL = '{TEAL}', AMBER = '{AMBER}';
const STATUS = {json.dumps(STATUS_COLORS, ensure_ascii=False)};
const regionRaw = {json.dumps(region_chart, ensure_ascii=False)};
const statusRaw = {json.dumps(status_chart, ensure_ascii=False)};
const progressRaw = {json.dumps(progress_chart, ensure_ascii=False)};
new Chart(document.getElementById('c1'), {{
  type: 'bar',
  data: {{ labels: regionRaw.map(d => d.name),
    datasets: [{{ data: regionRaw.map(d => d.value), backgroundColor: NAVY, borderRadius: 5 }}] }},
  options: {{ responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }},
      datalabels: {{ anchor:'end', align:'end', color: NAVY,
        font: {{ family:'Cairo', weight:'bold', size:11 }},
        formatter: v => v ? (v/1e6).toFixed(0)+'م' : '' }} }},
    scales: {{ x: {{ grid: {{ display:false }}, ticks: {{ font: {{ family:'Cairo' }} }} }},
              y: {{ beginAtZero:true, ticks: {{ callback: v => (v/1e6).toFixed(0)+'م' }}, grid: {{ color:'#f1f5f9' }} }} }},
    layout: {{ padding: {{ top: 18 }} }} }}
}});
new Chart(document.getElementById('c2'), {{
  type: 'bar',
  data: {{ labels: statusRaw.map(d => d.name),
    datasets: [{{ data: statusRaw.map(d => d.count),
      backgroundColor: statusRaw.map(d => STATUS[d.name] || TEAL), borderRadius: 5 }}] }},
  options: {{ responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }},
      datalabels: {{ anchor:'end', align:'end', color: NAVY,
        font: {{ family:'Cairo', weight:'bold', size:12 }}, formatter: v => v || '' }} }},
    scales: {{ x: {{ grid: {{ display:false }}, ticks: {{ font: {{ family:'Cairo' }} }} }},
              y: {{ beginAtZero:true, ticks: {{ stepSize: 1 }}, grid: {{ color:'#f1f5f9' }} }} }},
    layout: {{ padding: {{ top: 18 }} }} }}
}});
new Chart(document.getElementById('c3'), {{
  type: 'bar',
  data: {{ labels: progressRaw.map(d => d.name),
    datasets: [
      {{ label: 'مخطط', data: progressRaw.map(d => d.planned), backgroundColor: TEAL, borderRadius: 4 }},
      {{ label: 'فعلي', data: progressRaw.map(d => d.actual), backgroundColor: AMBER, borderRadius: 4 }}
    ] }},
  options: {{ responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ position:'top', labels: {{ font: {{ family:'Cairo' }} }} }},
      datalabels: {{ display: false }} }},
    scales: {{ x: {{ grid: {{ display:false }}, ticks: {{ font: {{ family:'Cairo' }} }} }},
              y: {{ beginAtZero:true, max: 110, ticks: {{ callback: v => v+'%' }}, grid: {{ color:'#f1f5f9' }} }} }} }}
}});
</script>
</body>
</html>
"""
    path = OUT / 'projects_dashboard.html'
    path.write_text(html, encoding='utf-8')
    return path


def main():
    p = argparse.ArgumentParser()
    p.add_argument('excel', nargs='?', default=None)
    p.add_argument('--demo', action='store_true')
    args = p.parse_args()
    demo = args.demo
    src = Path(args.excel) if args.excel else DEFAULT_XLSX
    if not demo and src.exists():
        rows = load_excel(src)
        print(f'Loaded {len(rows)} included projects from {src}')
    else:
        if not demo and not src.exists():
            print(f'Excel not found ({src}) — DEMO numbers for layout only')
        rows = demo_rows()
        demo = True
        print(f'DEMO rows: {len(rows)} (not production)')
    s = summarize(rows)
    charts = save_charts(s, demo)
    html = write_html(s, demo)
    print('Wrote:')
    print(f'  {html}')
    for c in charts:
        print(f'  {c}')


if __name__ == '__main__':
    main()