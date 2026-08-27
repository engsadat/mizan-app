# scripts/gen_org_chart.py
# Generates professional org chart HTML for each NWC southern region.
# Layout: A3 landscape pages (277mm height), card-based with proper page breaks.
# Data source: Excel files (same as HR/org_charts version)
#
# Outputs:
#   app/static/org_charts/09_OrgChart_Asir.html
#   app/static/org_charts/10_OrgChart_Jizan.html
#   app/static/org_charts/11_OrgChart_Baha.html
#   app/static/org_charts/12_OrgChart_Najran.html
#
# Run from mizan-app root: python scripts/gen_org_chart.py

import os, re, shutil, sys, base64
from pathlib import Path
from collections import defaultdict
from datetime import date
import openpyxl

_MONTHS_AR = ['يناير','فبراير','مارس','أبريل','مايو','يونيو',
               'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']
_t = date.today()
REPORT_MONTH = f"{_MONTHS_AR[_t.month - 1]} {_t.year}"

# MIZAN PATHS: script is in mizan-app/scripts/
BASE     = Path(__file__).parent.parent  # mizan-app root
DATA_DIR = BASE / 'data'  # Excel data folder (should be synced/deployed)
OUT_DIR  = BASE / 'app' / 'static' / 'org_charts'
OUT_DIR.mkdir(parents=True, exist_ok=True)

def _b64img(p):
    """Load image and convert to base64 data URL."""
    if not p.exists():
        print(f"Warning: Image not found: {p}")
        return None
    with open(p, 'rb') as f:
        return 'data:image/png;base64,' + base64.b64encode(f.read()).decode()

# Logo paths — adjust based on server structure
# Option 1: If logos are in app/static/images/
LOGO_PATH_BASE = BASE / 'app' / 'static' / 'images'
if not LOGO_PATH_BASE.exists():
    # Option 2: Try synced HR data folder
    LOGO_PATH_BASE = DATA_DIR / 'NWC layout' / 'img'

LOGO_ALAMRO = _b64img(LOGO_PATH_BASE / 'Alamro_Logo.png')
LOGO_NWC    = _b64img(LOGO_PATH_BASE / 'NWC_Logo.png')

if not LOGO_ALAMRO or not LOGO_NWC:
    print("WARNING: Logos not found. Using placeholder SVGs.")
    LOGO_ALAMRO = 'data:image/svg+xml,%3Csvg%3E%3Ctext%3EAl-Amro%3C/text%3E%3C/svg%3E'
    LOGO_NWC = 'data:image/svg+xml,%3Csvg%3E%3Ctext%3ENWC%3C/text%3E%3C/svg%3E'

ERR_VALS = {'#REF!', '=#REF!', '#VALUE!', '#N/A', '#NAME?', '#DIV/0!', '#NULL!', '#NUM!'}

def sv(v):
    """Safe value: convert to string, strip, return empty if error value."""
    if v is None: return ''
    s = str(v).strip()
    return '' if s in ERR_VALS else s

def load_copy(path):
    """Load Excel file safely (avoids lock if file is open)."""
    tmp = str(path) + '.tmp.xlsx'
    shutil.copy2(path, tmp)
    wb = openpyxl.load_workbook(tmp, data_only=True)
    os.remove(tmp)
    return wb

# Load exception prefixes for name shortening
_exc_path = DATA_DIR / 'Organize' / 'exception shorten names.txt'
EXCEPTION_PREFIXES = set()
if _exc_path.exists():
    for _ln in _exc_path.read_text(encoding='utf-8').splitlines():
        _w = _ln.strip()
        if _w:
            EXCEPTION_PREFIXES.add(_w)

def _name_units(name):
    """Split name into units, respecting exception prefixes."""
    parts = sv(name).split()
    units = []
    i = 0
    while i < len(parts):
        if parts[i] in EXCEPTION_PREFIXES and i + 1 < len(parts):
            units.append(parts[i] + ' ' + parts[i + 1])
            i += 2
        else:
            units.append(parts[i])
            i += 1
    return units

def short3(name): return ' '.join(_name_units(name)[:3])
def short2(name): return ' '.join(_name_units(name)[:2])

def fmt_val(v):
    """Format value in millions with م suffix."""
    try:
        n = float(str(v).replace(',', ''))
        if n <= 0: return ''
        m = n / 1_000_000
        return f'{m:.0f}م' if m >= 10 else f'{m:.1f}م'
    except:
        return ''

def fmt_val_sar(v):
    """Format total value for cover KPI — billions or millions with ر.س label."""
    try:
        n = float(str(v).replace(',', ''))
        if n <= 0: return ''
        if n >= 1_000_000_000:
            return f'{n / 1_000_000_000:.2f} مليار ر.س'
        return f'{n / 1_000_000:,.0f} مليون ر.س'
    except:
        return ''

def strip_grade(job):
    """Remove E2/E3 grade suffix from job title."""
    return re.sub(r'\s+E\d+$', '', sv(job)).strip()

RE_MAP  = {}
CON_MAP = {
    'شركة الكبريش للاستثمار و المقاولات':            'شركة الكبريش للاستثمار والمقاولات',
    'شركة الجازع للمقاولات والتجارة (مساهمة مقفلة)': 'شركة الجازع للمقاولات والتجارة',
    'شركة عزم للإستشمار':                            'شركة عزم للإستثمار',
    'شركة نبراس المتحدة للمقاولات العامة':            'شركة نبراس المتحدة للمقاولات',
    'شركة وتد العربية .':                            'شركة وتد العربية ذات مسئولية محدودة',
    'مؤسسة العرض الانشائي للمقاولات':                'شركة العرض الانشائي للمقاولات شركة شخص واحد',
    'مؤسسة سكاي العربية للمقاولات':                  'شركة سكاي العربية للمقاولات',
    'مؤسسة عبد المحسن شعبان القحطانى للمقاولات':     'شركة عبد المحسن شعبان القحطانى للمقاولات',
}

STATUS_CFG = {
    'تحت التنفيذ':                              ('#0071B9', '#fff'),
    'ايقاف كلي':                               ('#991B1B', '#fff'),
    'استلام ابتدائي(أثناء فترة الـ 60 يوم)':   ('#7C3AED', '#fff'),
    'استلام ابتدائي':                           ('#FFC000', '#1e293b'),
    'مسحوب':                                   ('#C00000', '#fff'),
    'إيقاف جزئي':                              ('#F59E0B', '#fff'),
}

def status_style(s):
    """Get color for status."""
    for k, (bg, fg) in STATUS_CFG.items():
        if k in sv(s):
            return bg, fg
    return '#72808A', '#fff'

REGION_COL = {'عسير': 30, 'جازان': 31, 'الباحة': 32, 'نجران': 33}
REGION_FILE = {
    'عسير':   '09_OrgChart_Asir.html',
    'جازان':  '10_OrgChart_Jizan.html',
    'الباحة': '11_OrgChart_Baha.html',
    'نجران':  '12_OrgChart_Najran.html',
}
REGION_GROUPS = {
    'عسير':   [4, 3, 3, 3],
    'جازان':  [3, 3],
    'الباحة': [2, 2],
    'نجران':  [3],
}

# ── 0. Job sort order ─────────────────────────────────────────────────────────
try:
    wb_sort = load_copy(DATA_DIR / 'Organize' / 'emp_sort.xlsx')
    ws_sort = wb_sort['Sheet1']
    JOB_SORT = {}
    for row in ws_sort.iter_rows(min_row=2, max_row=ws_sort.max_row, values_only=True):
        full_job = sv(row[0])
        rank     = row[1]
        if full_job and rank:
            stripped = re.sub(r'\s+E\d+$', '', full_job).strip()
            try:
                JOB_SORT.setdefault(stripped, int(rank))
            except (ValueError, TypeError):
                pass
    print(f"Loaded {len(JOB_SORT)} job sort orders")
except Exception as e:
    print(f"Warning: Could not load job sort: {e}")
    JOB_SORT = {}

# ── 1. Office-RE directory ────────────────────────────────────────────────────
try:
    wb_re = load_copy(DATA_DIR / 'Organize' / 'Office-RE.xlsx')
    ws_re = wb_re['RE_Mail']
    re_info   = {}
    reg_order = defaultdict(list)

    for row in ws_re.iter_rows(min_row=2, max_row=ws_re.max_row, values_only=True):
        name    = sv(row[1])
        region  = sv(row[4])
        ocode   = sv(row[7])
        oname   = sv(row[8])
        phone   = sv(row[3])
        page_no = sv(row[9])
        if not name or not region:
            continue
        if ocode == 'عسير_0':
            continue
        re_info[name] = {'phone': phone, 'office_code': ocode, 'office_name': oname, 'page_no': page_no}
        reg_order[region].append((ocode, name))

    def _ocode_key(ocode):
        parts = re.split(r'(\d+)', ocode)
        return [int(p) if p.isdigit() else p for p in parts]

    for region in reg_order:
        reg_order[region].sort(key=lambda x: _ocode_key(x[0]))
    print(f"Loaded RE directory: {sum(len(v) for v in reg_order.values())} REs")
except Exception as e:
    print(f"Warning: Could not load RE directory: {e}")
    re_info = {}
    reg_order = defaultdict(list)

# ── 2. Employees ──────────────────────────────────────────────────────────────
try:
    wb_emp = load_copy(DATA_DIR / 'source' / 'employees data source.xlsx')
    ws_emp = wb_emp['data']
    ws_stk = wb_emp['Stackholders']
    emp_by_re = defaultdict(list)
    region_emp_total = defaultdict(int)

    for row in ws_emp.iter_rows(min_row=2, max_row=ws_emp.max_row, values_only=True):
        if sv(row[11]) != 'على قوة العمل':
            continue
        re_name  = sv(row[27])
        emp_name = sv(row[20])
        job      = strip_grade(row[21])
        nation   = sv(row[14])
        region   = sv(row[16])
        if region in {'عسير', 'جازان', 'الباحة', 'نجران'}:
            region_emp_total[region] += 1
        if re_name and emp_name:
            emp_by_re[re_name].append({'name': emp_name, 'job': job, 'nation': nation})

    ws_cat = wb_emp['emp_Category']
    JOB_CATS       = []
    _EMP_CAT_LOOKUP = {}
    for row in ws_cat.iter_rows(min_row=2, max_row=ws_cat.max_row, values_only=True):
        raw = sv(row[1])
        if not raw:
            continue
        cat = strip_grade(raw)
        _EMP_CAT_LOOKUP[raw] = cat
        _EMP_CAT_LOOKUP[cat] = cat
        if cat not in JOB_CATS:
            JOB_CATS.append(cat)

    total_emp = sum(len(e) for e in emp_by_re.values())
    print(f"Loaded {total_emp} employees")
except Exception as e:
    print(f"Error loading employees: {e}")
    emp_by_re = defaultdict(list)
    region_emp_total = defaultdict(int)
    JOB_CATS = []
    _EMP_CAT_LOOKUP = {}

# ── 2c. Specialists ───────────────────────────────────────────────────────────
try:
    wb_spec = load_copy(DATA_DIR / 'source' / 'employees data source.xlsx')
    ws_spec = wb_spec['ميكانيكا- كهرباء']
    VALID_REGIONS = {'عسير', 'جازان', 'الباحة', 'نجران'}
    GENERAL_RE    = {'All', 'السلامة', ''}

    def _spec_cat(job):
        j = sv(job)
        if 'ميكانيكا' in j: return 'mek'
        if 'كهرباء'   in j: return 'elec'
        if 'أمن' in j or 'سلامة' in j: return 'safe'
        return None

    spec_by_region = defaultdict(list)
    for row in ws_spec.iter_rows(min_row=2, values_only=True):
        sup  = sv(row[1])
        reg  = sv(row[2])
        name = sv(row[3])
        job  = sv(row[4])
        re1  = sv(row[5])
        re2  = sv(row[6])
        if not name or sup == 'دعم فني':
            continue
        cat = _spec_cat(job)
        if not cat or reg not in VALID_REGIONS:
            continue
        spec_by_region[reg].append({'name': name, 'cat': cat, 're1': re1, 're2': re2})

    spec_names = {sp['name'] for lst in spec_by_region.values() for sp in lst}
    print(f"Loaded {sum(len(v) for v in spec_by_region.values())} specialists")
except Exception as e:
    print(f"Warning: Could not load specialists: {e}")
    spec_by_region = defaultdict(list)
    spec_names = set()

# ── 2d. Stakeholders ──────────────────────────────────────────────────────────
try:
    fixed_stakeholders = []
    region_stakeholder = {}
    for _row in ws_stk.iter_rows(min_row=2, max_row=ws_stk.max_row, values_only=True):
        _role = sv(_row[0])
        _name = sv(_row[1])
        _reg  = sv(_row[2])
        if not _role or not _name or not _reg:
            continue
        if _reg == 'All':
            fixed_stakeholders.append((_role, _name))
        else:
            region_stakeholder[_reg] = (_role, _name)
    print(f"Loaded {len(fixed_stakeholders)} fixed stakeholders")
except Exception as e:
    print(f"Warning: Could not load stakeholders: {e}")
    fixed_stakeholders = []
    region_stakeholder = {}

# ── 3. Projects ───────────────────────────────────────────────────────────────
try:
    wb_pro = load_copy(DATA_DIR / 'source' / 'project_2026_database_ver1_updated.xlsx')
    ws_pro = wb_pro['pro']
    STATUS_RANK = {}
    for _row in wb_pro['Status'].iter_rows(min_row=2, values_only=True):
        _sname = sv(_row[0])
        _rank  = _row[1]
        if _sname and _rank is not None:
            try: STATUS_RANK[_sname] = int(_rank)
            except: pass

    proj_by_re = {r: defaultdict(list) for r in REGION_COL}
    for row in ws_pro.iter_rows(min_row=2, max_row=ws_pro.max_row, values_only=True):
        if not (row[10] and str(row[10]).lower() == 'yes'):
            continue
        proj_name  = sv(row[12])
        contractor = CON_MAP.get(sv(row[18]), sv(row[18]))
        status     = sv(row[23])
        value      = row[29]
        for region, cidx in REGION_COL.items():
            re_name = RE_MAP.get(sv(row[cidx]), sv(row[cidx]))
            if re_name:
                proj_by_re[region][re_name].append({
                    'name': proj_name, 'contractor': contractor,
                    'status': status,  'value': value,
                    'type': sv(row[28]),
                })

    total_proj = sum(len(v) for d in proj_by_re.values() for v in d.values())
    print(f"Loaded {total_proj} projects")
except Exception as e:
    print(f"Error loading projects: {e}")
    proj_by_re = {r: defaultdict(list) for r in REGION_COL}
    STATUS_RANK = {}

# ── 4. CSS ────────────────────────────────────────────────────────────────────
CSS = '''
:root {
  --navy:#002060; --blue:#0071B9; --sky:#00A7E2; --teal:#03A88B;
  --amber:#d97706; --red:#C00000; --gray:#72808A;
  --lb:#e8f4fd; --lg:#f5f7fa; --bd:#d1d9e6;
}
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html { width: 392mm; }
body { font-family: 'Cairo', sans-serif; direction: rtl; font-size: 11px; color: var(--navy); background: #fff; width: 100%; }

/* ── ONE PAGE = ONE A3 LANDSCAPE SHEET ─────────────────────────── */
/* Content area: 420mm wide - 2x14mm margins = 392mm = 1481px      */
/*               297mm tall - 2x10mm margins = 277mm               */
.page {
  width: 100%;
  height: 277mm;
  display: flex;
  flex-direction: column;
  overflow: hidden;
  break-after: page;
}
.page:last-child { break-after: auto; }

/* ── HEADER ──────────────────────────────────────────────────────── */
.ph {
  display: flex; align-items: center; gap: 8px; padding: 3px 10px;
  border-bottom: 3px solid var(--blue); border-top: 3px solid var(--amber);
  flex-shrink: 0;
}
.ph img { height: 26px; max-width: 90px; object-fit: contain; flex-shrink: 0; }
.ph-mid { flex: 1; text-align: center; }
.ph-mid h1 { font-size: 13px; font-weight: 900; line-height: 1.3; }
.ph-mid small { font-size: 9px; color: var(--gray); }

.pmgr {
  background: var(--lb); padding: 2px 10px;
  display: flex; align-items: center; justify-content: center; gap: 10px;
  border-bottom: 1px solid var(--bd); flex-shrink: 0;
}
.mi { font-size: 9.5px; font-weight: 700; display: flex; align-items: center; gap: 4px; }
.ml { color: var(--gray); font-weight: 400; font-size: 8.5px; }
.ms { width: 1px; height: 14px; background: var(--bd); }

/* ── CARDS ROW — fills height between header and footer ─────────── */
.cards {
  flex: 1; min-height: 0;
  display: flex; gap: 2px; padding: 2px 0;
}

/* ── SINGLE CARD ─────────────────────────────────────────────────── */
.card {
  flex: 1; min-width: 0;
  display: flex; flex-direction: column;
  border: 1px solid var(--bd); border-top: 3px solid var(--blue);
  border-radius: 4px; overflow: hidden;
}

/* Office name header */
.coh {
  background: var(--lb); padding: 4px 8px;
  border-bottom: 1px solid var(--bd);
  display: flex; align-items: center; justify-content: center;
  gap: 5px; flex-wrap: wrap; flex-shrink: 0;
}
.con { font-size: 11px; font-weight: 900; line-height: 1.2; }
.oh-sep { font-size: 9px; color: var(--gray); }
.cnt { background: var(--blue); color: #fff; border-radius: 8px; padding: 0 5px; font-size: 8.5px; font-weight: 700; }

/* Card body — flex column, projects float top, RE+team anchor to bottom */
.cbody {
  flex: 1; min-height: 0;
  display: flex; flex-direction: column;
  padding: 3px; gap: 5px; overflow: hidden;
  justify-content: space-between;
}

/* Projects */
.pl { display: flex; flex-direction: column; gap: 4px; }
.pl.two-col { display: grid; grid-template-columns: 1fr 1fr; gap: 4px; }
.pi {
  background: #fff; border-radius: 4px; padding: 4px 7px;
  border: 1px solid var(--bd); border-right: 4px solid var(--blue);
  min-width: 0; overflow: hidden;
}
.pn { font-size: 11px; font-weight: 900; line-height: 1.3; margin-bottom: 2px; overflow: hidden; word-break: break-word; }
.pm { display: flex; align-items: center; justify-content: space-between; gap: 3px; flex-wrap: wrap; }
.pc { font-size: 8.5px; color: var(--gray); flex: 1; min-width: 0; overflow: hidden; word-break: break-word; }
.pb { display: flex; align-items: center; gap: 3px; flex-shrink: 0; }
.ps { font-size: 8px; font-weight: 700; padding: 1px 4px; border-radius: 3px; white-space: nowrap; border: 1px solid currentColor; }
.pv { font-size: 9px; font-weight: 900; color: var(--teal); }

/* Spacer: pushes RE strip and team to the bottom */
.spacer { flex: 1; min-height: 0; }

/* RE strip */
.re-strip {
  background: var(--lg); border: 1px solid var(--bd); border-radius: 3px;
  padding: 3px 6px; display: flex; align-items: center; justify-content: center;
  gap: 5px; flex-wrap: wrap; flex-shrink: 0;
}
.re-lbl { font-size: 8px; font-weight: 700; color: var(--gray); }
.re-nm  { font-size: 9.5px; font-weight: 900; }
.re-ph  { font-size: 8.5px; color: var(--blue); }

/* Team grid */
.team { display: grid; grid-template-columns: 1fr 1fr; gap: 2px; flex-shrink: 0; }
.eg { border-radius: 3px; border: 1px solid var(--bd); border-right: 3px solid var(--sky); overflow: hidden; }
.ej { font-size: 9px; font-weight: 900; background: var(--lb); padding: 2px 5px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.ec { font-size: 8px; color: var(--blue); background: #fff; border-radius: 8px; padding: 0 3px; border: 1px solid var(--bd); }
.ens { padding: 1px 3px; display: grid; gap: 1px; background: #fff; }
.en  { font-size: 8.5px; padding: 1px 3px; border-bottom: 1px solid var(--lg); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.en:last-child { border-bottom: none; }
.em  { font-size: 8.5px; color: var(--gray); text-align: center; padding: 3px; }

/* ── SPECIALIST FOOTER ROW ───────────────────────────────────────── */
.spec-area { display: flex; gap: 2px; padding: 0 0 2px; flex-shrink: 0; }
.spec-slot { flex: 1; }
.sftr { background: var(--lg); border-radius: 3px; border: 1px solid var(--bd); }
.sf-title { font-size: 9px; font-weight: 900; padding: 2px 6px 2px; border-bottom: 1px solid var(--bd); text-align: center; }
.srow { display: flex; align-items: center; gap: 6px; padding: 2px 5px; border-right: 3px solid; }
.srow.mek  { border-right-color: var(--amber); background: #fff8ed; }
.srow.elec { border-right-color: var(--sky);   background: #e8f8fd; }
.srow.safe { border-right-color: var(--teal);  background: #eafaf7; }
.slbl { font-size: 8px; font-weight: 900; white-space: nowrap; min-width: 55px; }
.srow.mek  .slbl { color: var(--amber); }
.srow.elec .slbl { color: var(--sky);   }
.srow.safe .slbl { color: var(--teal);  }
.schips { display: flex; flex-wrap: wrap; gap: 2px 8px; }
.schi   { font-size: 8px; display: inline-flex; align-items: center; gap: 3px; }
.snm    { font-weight: 700; color: var(--navy); }
.sof    { font-size: 7.5px; color: var(--gray); }

/* ── PAGE FOOTER ─────────────────────────────────────────────────── */
.pf {
  background: var(--lg); border-top: 2px solid var(--bd);
  color: var(--gray); font-size: 9px; padding: 3px 10px;
  display: flex; justify-content: space-between; align-items: center;
  flex-shrink: 0;
}

/* ── COVER PAGE ──────────────────────────────────────────────────── */
.cv-kpis {
  display: flex; gap: 8px; padding: 8px 0 6px; flex-shrink: 0;
}
.cv-kpi {
  flex: 1; background: #fff; border: 1px solid var(--bd);
  border-top: 4px solid var(--blue); border-radius: 6px;
  padding: 10px 8px; text-align: center;
}
.cv-kpi-num { font-size: 38px; font-weight: 900; color: var(--blue); line-height: 1; }
.cv-kpi-lbl { font-size: 11px; color: var(--gray); margin-top: 4px; }
.cv-kpi-split { display: flex; gap: 6px; justify-content: center; margin: 2px 0; flex-wrap: wrap; }
.cv-kpi-s1 { font-size: 9px; font-weight: 700; color: var(--teal); background: #eafaf7; border-radius: 10px; padding: 1px 7px; white-space: nowrap; }
.cv-kpi-s2 { font-size: 9px; font-weight: 700; color: var(--amber); background: #fff8e7; border-radius: 10px; padding: 1px 7px; white-space: nowrap; }

.cv-body { flex: 1; min-height: 0; display: flex; flex-direction: column; gap: 6px; padding-bottom: 6px; }

.cv-sec {
  flex: 1; display: flex; flex-direction: column; gap: 0;
  background: var(--lg); border: 1px solid var(--bd); border-radius: 6px;
  padding: 8px 10px; overflow: hidden;
}
.cv-sec-title {
  font-size: 12px; font-weight: 900; color: var(--navy);
  border-bottom: 2px solid var(--blue); padding-bottom: 5px; margin-bottom: 6px;
  flex-shrink: 0;
}
.cv-sec-sub {
  font-size: 9.5px; font-weight: 900; color: var(--gray);
  border-bottom: 1px solid var(--bd); padding-bottom: 2px;
  margin: 8px 0 5px; flex-shrink: 0;
}

/* Status / type bars */
.cv-bar-row { display: flex; align-items: center; gap: 6px; margin-bottom: 3px; flex-shrink: 0; }
.cv-bar-lbl { font-size: 8.5px; color: var(--navy); flex: 1; text-align: right; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.cv-bar     { width: 90px; height: 13px; background: var(--bd); border-radius: 3px; flex-shrink: 0; position: relative; overflow: hidden; }
.cv-bar-fill { height: 100%; border-radius: 3px; }
.cv-bar-n   { font-size: 8px; font-weight: 900; color: var(--navy); flex-shrink: 0; min-width: 18px; text-align: left; }

/* RE detail grid — one card per RE office */
.cv-re-grid { display: grid; gap: 6px; overflow: hidden; }
.cv-re-card {
  background: #fff; border: 1px solid var(--bd); border-top: 4px solid var(--blue);
  border-radius: 5px; padding: 8px 10px; display: flex; flex-direction: column;
  gap: 0; justify-content: space-evenly; overflow: hidden;
}
.cv-re-coh  { font-size: 12px; font-weight: 900; color: var(--navy); border-bottom: 1px solid var(--bd); padding-bottom: 4px; margin-bottom: 4px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; text-align: center; }
.cv-re-re   { color: var(--blue); }
.cv-re-line { display: flex; align-items: center; gap: 5px; flex-wrap: wrap; padding: 2px 0; }
.cv-re-ttl  { font-size: 11px; font-weight: 900; white-space: nowrap; flex-shrink: 0; }
.cv-re-ttl.proj { color: var(--blue); }
.cv-re-ttl.emp  { color: var(--teal); }
.cv-re-ttl.con  { color: var(--amber); }
.cv-chip    { font-size: 9.5px; font-weight: 700; border-radius: 4px; padding: 1px 6px; border: 1px solid currentColor; white-space: nowrap; }
.cv-jchip   { font-size: 9.5px; background: var(--lg); border-radius: 4px; padding: 1px 6px; color: var(--navy); white-space: nowrap; }

/* Employee classification rows */
.cv-emp-row  { display: flex; align-items: center; gap: 6px; margin-bottom: 3px; flex-shrink: 0; }
.cv-emp-lbl  { font-size: 8.5px; color: var(--navy); flex: 1; }
.cv-emp-n    { font-size: 8px; font-weight: 900; color: var(--teal); background: #eafaf7; border-radius: 8px; padding: 0 6px; flex-shrink: 0; }

/* Nationality chips */
.cv-nat { display: flex; gap: 8px; flex-wrap: wrap; margin-top: 4px; }
.cv-nat-chip { display: flex; align-items: center; gap: 4px; border-radius: 6px; padding: 3px 10px; flex: 1; justify-content: center; }
.cv-nat-chip.sa  { background: #e8f4fd; border: 1px solid var(--blue); }
.cv-nat-chip.nsa { background: var(--lg); border: 1px solid var(--bd); }
.cv-nat-n   { font-size: 18px; font-weight: 900; }
.cv-nat-chip.sa  .cv-nat-n { color: var(--blue); }
.cv-nat-chip.nsa .cv-nat-n { color: var(--gray); }
.cv-nat-lbl { font-size: 8.5px; color: var(--gray); }

/* Job summary strip */
.cv-job-bar {
  display: flex; align-items: center; flex-wrap: wrap; gap: 6px;
  background: #f0f4fa; border: 1px solid var(--bd); border-radius: 5px;
  padding: 6px 10px; flex-shrink: 0;
}
.cv-job-lbl { font-size: 9.5px; font-weight: 900; color: var(--gray); flex-shrink: 0; white-space: nowrap; margin-left: 4px; }
.cv-jbar-chip {
  font-size: 10px; font-weight: 700; background: #fff; border: 1px solid var(--bd);
  border-radius: 4px; padding: 2px 8px; color: var(--navy); white-space: nowrap;
  display: inline-flex; align-items: center; gap: 4px;
}
.cv-jbar-chip .jn { font-size: 11px; font-weight: 900; color: var(--blue); }

/* Signature block */
.cv-sigs { display: flex; flex-direction: row; gap: 10px; align-items: stretch; flex-shrink: 0; border-top: 2px solid var(--bd); padding-top: 8px; padding-bottom: 4px; }
.cv-sig {
  flex: 1; border: 1px solid var(--bd); border-top: 3px solid var(--teal);
  border-radius: 6px; padding: 7px 16px 9px;
  display: flex; flex-direction: column; align-items: center; gap: 2px;
  background: #f8fffe;
}
.cv-sig-head {
  flex: 1; margin-top: 5px;
  border: 2px solid var(--navy); border-top: 5px solid var(--amber);
  background: #fff;
}
.cv-sig-role           { font-size: 8.5px; color: var(--gray); }
.cv-sig-head .cv-sig-role { font-size: 9px; font-weight: 700; color: var(--amber); }
.cv-sig-name           { font-size: 11px; font-weight: 900; color: var(--navy); margin-bottom: 16px; }
.cv-sig-head .cv-sig-name { font-size: 13px; color: var(--navy); }
.cv-sig-line           { width: 75%; height: 1px; background: var(--teal); margin-bottom: 3px; }
.cv-sig-head .cv-sig-line { background: var(--amber); }
.cv-sig-tag            { font-size: 8px; color: var(--gray); }
.cv-sig-head .cv-sig-tag  { color: var(--gray); }

@page { size: A3 landscape; margin: 10mm 14mm; }
@media print {
  html { overflow: hidden; }
  body { background: #fff; }
  .page { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
}
'''

# ── 5. Card HTML builder ──────────────────────────────────────────────────────
def card_html(re_name, info, projects, employees):
    projs = sorted(projects, key=lambda p: STATUS_RANK.get(p['status'], 99))
    proj_html = ''
    for p in projs:
        bg, fg = status_style(p['status'])
        val     = fmt_val(p['value'])
        val_h   = f'<span class="pv">{val}</span>' if val else ''
        proj_html += (
            f'<div class="pi" style="border-right-color:{bg}">'
            f'<div class="pn">{p["name"]}</div>'
            f'<div class="pm">'
            f'<span class="pc">{p["contractor"]}</span>'
            f'<span class="pb">'
            f'<span class="ps" style="color:{bg};border-color:{bg}">{p["status"]}</span>'
            f'{val_h}</span></div></div>'
        )

    job_groups = {}
    for e in employees:
        if e['job'] == 'مهندس مقيم': continue
        if e['name'] in spec_names:   continue
        job_groups.setdefault(e['job'], []).append(short3(e['name']))
    team_html = ''
    for job, names in sorted(job_groups.items(), key=lambda x: JOB_SORT.get(x[0], 9999)):
        names_h = ''.join(f'<div class="en">{n}</div>' for n in names)
        ncols   = min(len(names), 2)
        team_html += (
            f'<div class="eg">'
            f'<div class="ej">{job} <span class="ec">{len(names)}</span></div>'
            f'<div class="ens" style="grid-template-columns:repeat({ncols},1fr)">{names_h}</div>'
            f'</div>'
        )

    n_team  = sum(1 for e in employees if e['job'] != 'مهندس مقيم' and e['name'] not in spec_names)
    oname   = info.get('office_name', re_name)
    phone   = info.get('phone', '')
    two_col = ' two-col' if len(projs) >= 6 else ''

    return (
        f'<div class="card">'
        f'<div class="coh">'
        f'<span class="con">{oname}</span>'
        f'<span class="oh-sep">|</span>'
        f'<span class="cnt">📋 {len(projs)}</span>'
        f'</div>'
        f'<div class="cbody">'
        f'<div class="pl{two_col}">'
        f'{proj_html or "<div class=\"em\">—</div>"}'
        f'</div>'
        f'<div class="re-strip">'
        f'<span class="re-lbl">المهندس المقيم</span>'
        f'<span class="oh-sep">|</span>'
        f'<span class="re-nm">{re_name}</span>'
        + (f'<span class="oh-sep">|</span><span class="re-ph">📞 {phone}</span>' if phone else '')
        + f'<span class="oh-sep">|</span>'
        f'<span class="re-lbl">👥</span><span class="cnt">{n_team}</span>'
        f'</div>'
        f'<div class="team">{team_html or "<div class=\"em\">—</div>"}</div>'
        f'</div>'
        f'</div>'
    )

# ── 6. Specialist slot builder ────────────────────────────────────────────────
CATS = [('mek', 'ميكانيكا'), ('elec', 'كهرباء'), ('safe', 'أمن وسلامة')]

def slot_html(specs, show_office=False):
    rows = ''
    for cat_id, cat_lbl in CATS:
        items = [sp for sp in specs if sp['cat'] == cat_id]
        if not items: continue
        chips = ''
        for sp in items:
            chips += f'<span class="schi"><span class="snm">{short3(sp["name"])}</span>'
            if show_office:
                offices = []
                for rn in (sp['re1'], sp['re2']):
                    if rn and rn not in GENERAL_RE:
                        o = re_info.get(rn, {}).get('office_name', '') or rn
                        if o: offices.append(o)
                if offices:
                    chips += f'<span class="sof">({" · ".join(offices)})</span>'
            chips += '</span>'
        rows += (f'<div class="srow {cat_id}">'
                 f'<span class="slbl">{cat_lbl}</span>'
                 f'<div class="schips">{chips}</div></div>')
    if not rows: return ''
    return f'<div class="sftr">{rows}</div>'

def spec_area_html(page_re_names, region_specs):
    ncols     = len(page_re_names)
    re_to_col = {rn: i for i, rn in enumerate(page_re_names)}

    col_specs = defaultdict(list)
    all_specs = []
    seen_col  = defaultdict(set)
    seen_all  = set()

    for sp in region_specs:
        gen = sp['re1'] in GENERAL_RE and not sp['re2']
        if gen:
            if sp['name'] not in seen_all:
                seen_all.add(sp['name'])
                all_specs.append(sp)
            continue
        c1 = re_to_col.get(sp['re1']) if sp['re1'] not in GENERAL_RE else None
        c2 = re_to_col.get(sp['re2']) if sp['re2'] and sp['re2'] not in GENERAL_RE else None
        for c in (c1, c2):
            if c is not None and sp['name'] not in seen_col[c]:
                seen_col[c].add(sp['name'])
                col_specs[c].append(sp)

    if not col_specs and not all_specs:
        return ''

    parts = [f'<div class="sf-title">⚙️ وظائف عامة</div>']

    if col_specs:
        slots = ''
        for col in range(ncols):
            specs = col_specs.get(col, [])
            inner = slot_html(specs) if specs else ''
            slots += f'<div class="spec-slot">{inner}</div>'
        parts.append(f'<div class="spec-area">{slots}</div>')

    if all_specs:
        inner = slot_html(all_specs, show_office=True)
        parts.append(f'<div class="spec-area"><div class="spec-slot" style="flex:none;width:100%">{inner}</div></div>')

    return '\n'.join(parts)

# ── 7. Signature block helper ─────────────────────────────────────────────────
def _build_sig_block(signatories):
    _STYLES = [
        ('border-top:4px solid #0071B9;border-color:#b8d4ef;background:#f0f7ff', '#0071B9'),
        ('border-top:4px solid #03A88B;border-color:#b3e8dd;background:#f0fdf9', '#03A88B'),
        ('border-top:4px solid #d97706;border-color:#fde68a;background:#fffbef', '#d97706'),
    ]

    def _sig_box(role, name, idx):
        box_style, line_col = _STYLES[idx % len(_STYLES)]
        return (
            f'<div class="cv-sig" style="{box_style}">'
            f'<div class="cv-sig-role">{role}</div>'
            f'<div class="cv-sig-name">{name}</div>'
            f'<div class="cv-sig-line" style="background:{line_col}"></div>'
            f'<div class="cv-sig-tag">التوقيع</div>'
            f'</div>'
        )

    proj_mgr   = None
    other_sigs = []
    for role, name in signatories:
        if 'الجنوبي' in role:
            proj_mgr = (role, name)
        else:
            other_sigs.append((role, name))

    ordered = list(other_sigs)
    if proj_mgr:
        ordered.append(proj_mgr)

    boxes = [_sig_box(r, n, i) for i, (r, n) in enumerate(ordered)]
    return f'<div class="cv-sigs">{"".join(boxes)}</div>'

# ── 8. Cover page ─────────────────────────────────────────────────────────────
def emp_job_cat(job):
    j = job
    if j in _EMP_CAT_LOOKUP:                              return _EMP_CAT_LOOKUP[j]
    if 'مدير المشروع' in j:                              return 'مدير المشروع'
    if 'مهندس' in j and ('أمن' in j or 'سلامة' in j):   return 'مهندس أمن وسلامة'
    if 'مراقب' in j and ('أمن' in j or 'سلامة' in j):   return 'مراقب أمن وسلامة'
    if 'مراقب موقع' in j:                                return 'مراقب موقع'
    if 'مهندس مقيم' in j:                                return 'مهندس مقيم'
    if 'مهندس موقع' in j:                                return 'مهندس موقع'
    if 'مواد' in j:    return 'مهندس مواد وجودة'
    if 'كميات' in j:   return 'حاسب كميات'
    if 'مساح' in j:    return 'مساح'
    if 'وثائق' in j:   return 'أخصائي وثائق'
    if 'تخطيط' in j:   return 'مهندس تخطيط'
    if 'كهرباء' in j:  return 'مهندس كهرباء'
    if 'ميكانيك' in j: return 'مهندس ميكانيكا'
    return j or 'أخرى'

def build_cover(region, all_re_names, proj_re, emp_re, mgr_html, signatories, tech_sup_count=0):
    all_projects  = [p for rn in all_re_names for p in proj_re.get(rn, [])]
    all_employees = [e for rn in all_re_names for e in emp_re.get(rn, [])]

    total_proj       = len(all_projects)
    field_emp_count  = len(all_employees)
    total_emp        = field_emp_count + tech_sup_count
    total_con        = len({p['contractor'] for p in all_projects if p['contractor']})
    total_off        = len(all_re_names)
    val_sum = 0
    for p in all_projects:
        try:
            val_sum += float(str(p['value']).replace(',', ''))
        except: pass
    val_str = fmt_val_sar(val_sum)

    emp_split = (
        f'<div class="cv-kpi-split">'
        f'<span class="cv-kpi-s1">{field_emp_count} إشراف</span>'
        f'<span class="cv-kpi-s2">{tech_sup_count} دعم تقني</span>'
        f'</div>'
    ) if tech_sup_count else ''

    kpi_cards = (
        f'<div class="cv-kpi"><div class="cv-kpi-num">{total_proj}</div><div class="cv-kpi-lbl">مشروع نشط</div></div>'
        f'<div class="cv-kpi"><div class="cv-kpi-num">{total_con}</div><div class="cv-kpi-lbl">مقاول</div></div>'
        f'<div class="cv-kpi"><div class="cv-kpi-num">{total_emp}</div>{emp_split}<div class="cv-kpi-lbl">موظف</div></div>'
        f'<div class="cv-kpi"><div class="cv-kpi-num">{total_off}</div><div class="cv-kpi-lbl">مكتب إشراف</div></div>'
        + (f'<div class="cv-kpi"><div class="cv-kpi-num" style="font-size:22px">{val_str}</div><div class="cv-kpi-lbl">إجمالي العقود</div></div>' if val_str else '')
    )

    from collections import Counter
    JOB_SHORT_BAR = {
        'مدير المشروع':      'م.مشروع',
        'مهندس مقيم':        'م.مقيم',
        'مهندس موقع':        'م.موقع',
        'مراقب موقع':        'مراقب',
        'مهندس مواد وجودة':  'م.مواد',
        'حاسب كميات':        'ح.كميات',
        'مساح':              'مساح',
        'مراقب أمن وسلامة':  'م.سلامة',
        'مهندس أمن وسلامة':  'مه.سلامة',
        'أخصائي وثائق':      'أخ.وثائق',
        'مهندس تخطيط':       'م.تخطيط',
        'مهندس كهرباء':      'م.كهرباء',
        'مهندس ميكانيكا':    'م.ميكا',
    }
    bar_counts_full = Counter(emp_job_cat(e['job']) for e in all_employees if e['job'])
    if 'مهندس مقيم' in bar_counts_full:
        bar_counts_full['مهندس مقيم'] = total_off
    JOB_BAR_ORDER = [c for c in JOB_CATS if c in JOB_SHORT_BAR]
    job_bar_chips = ''
    for cat, cnt in sorted(bar_counts_full.items(),
                           key=lambda x: JOB_BAR_ORDER.index(x[0]) if x[0] in JOB_BAR_ORDER else 99):
        job_bar_chips += (
            f'<span class="cv-jbar-chip">'
            f'{cat} <span class="jn" dir="ltr">{cnt}</span>'
            f'</span>'
        )
    job_bar_html = (
        f'<div class="cv-job-bar">'
        f'<span class="cv-job-lbl">👥 التخصصات:</span>'
        f'{job_bar_chips}'
        f'</div>'
    )

    n_re = len(all_re_names)
    cols = 4 if n_re >= 10 else (3 if n_re >= 4 else 2)

    re_cards_html = ''
    for rn in all_re_names:
        oname  = re_info.get(rn, {}).get('office_name', rn)
        projs  = proj_re.get(rn, [])
        emps   = emp_re.get(rn, [])
        n_proj = len(projs)
        n_emp  = len(emps)
        n_con  = len({p['contractor'] for p in projs if p['contractor']})

        STATUS_LABEL = {
            'استلام ابتدائي(أثناء فترة الـ 60 يوم)': 'استلام (60 يوم)',
        }
        st_counts = Counter(p['status'] for p in projs if p['status'])
        proj_chips = ''
        for st, cnt in sorted(st_counts.items(), key=lambda x: STATUS_RANK.get(x[0], 99)):
            bg, _ = status_style(st)
            lbl = STATUS_LABEL.get(st, st)
            proj_chips += f'<span class="cv-chip" style="color:{bg};border-color:{bg}">{lbl} <span dir="ltr">{cnt}</span></span>'

        job_counts_full = Counter(emp_job_cat(e['job']) for e in emps if e['job'])
        if 'مهندس مقيم' in job_counts_full:
            job_counts_full['مهندس مقيم'] = 1
        emp_chips = ''
        for cat, cnt in sorted(job_counts_full.items(),
                               key=lambda x: JOB_CATS.index(x[0]) if x[0] in JOB_CATS else 99):
            emp_chips += f'<span class="cv-jchip">{cat} <span dir="ltr">{cnt}</span></span>'

        ty_counts = Counter()
        for p in projs:
            ty = p.get('type', '')
            if 'مياه' in ty:      ty_counts['مياه'] += 1
            elif 'صرف' in ty:     ty_counts['صرف']  += 1
            elif ty:              ty_counts['أخرى'] += 1
        TY_ORDER = {'مياه': 0, 'صرف': 1, 'أخرى': 2}
        ty_chips = ''
        for ty, cnt in sorted(ty_counts.items(), key=lambda x: TY_ORDER.get(x[0], 9)):
            col = 'var(--blue)' if ty == 'مياه' else ('var(--teal)' if ty == 'صرف' else 'var(--gray)')
            ty_chips += f'<span class="cv-chip" style="color:{col};border-color:{col}">{ty} <span dir="ltr">{cnt}</span></span>'

        re_cards_html += (
            f'<div class="cv-re-card">'
            f'<div class="cv-re-coh">{oname}&nbsp;&nbsp;&nbsp;<span class="cv-re-re">م/ {rn}</span></div>'
            f'<div class="cv-re-line">'
            f'<span class="cv-re-ttl proj">📋 {n_proj}</span>'
            f'{proj_chips}'
            f'</div>'
            f'<div class="cv-re-line">'
            f'<span class="cv-re-ttl emp">👥 {n_emp}</span>'
            f'{emp_chips}'
            f'</div>'
            f'<div class="cv-re-line">'
            f'<span class="cv-re-ttl con">مقاول <span dir="ltr">{n_con}</span></span>'
            f'{ty_chips}'
            f'</div>'
            f'</div>'
        )

    sig_html = _build_sig_block(signatories)

    return (
        f'<div class="page">'
        f'<div class="ph">'
        f'<img src="{LOGO_ALAMRO}" alt="Al-Amro">'
        f'<div class="ph-mid">'
        f'<h1>الهيكل التنظيمي — منطقة {region}</h1>'
        f'</div>'
        f'<img src="{LOGO_NWC}" alt="NWC">'
        f'</div>'
        f'<div class="pmgr">{mgr_html}</div>'
        f'<div class="cv-kpis">{kpi_cards}</div>'
        f'{job_bar_html}'
        f'<div class="cv-body">'
        f'<div class="cv-sec-title" style="flex-shrink:0">📍 ملخص مكاتب الإشراف ({total_off} مكتب) — المشاريع والموظفون والمقاولون</div>'
        f'<div class="cv-re-grid" style="grid-template-columns:repeat({cols},1fr);grid-auto-rows:1fr;flex:1;min-height:0;">'
        f'{re_cards_html}'
        f'</div>'
        f'</div>'
        f'{sig_html}'
        f'<div class="pf">'
        f'<span>شركة مجموعة العمرو للاستشارات الهندسية</span>'
        f'<span>شركة المياه الوطنية — المنطقة الجنوبية</span>'
        f'<span>صفحة الغلاف — {REPORT_MONTH}</span>'
        f'</div>'
        f'</div>'
    )

# ── 9. Page HTML builder ──────────────────────────────────────────────────────
def build_page(region, page_num, total_pages, re_names, proj_re, emp_re,
               mgr_html, region_specs, signatories):
    hdr = (
        f'<div class="ph">'
        f'<img src="{LOGO_ALAMRO}" alt="Al-Amro">'
        f'<div class="ph-mid">'
        f'<h1>الهيكل التنظيمي — منطقة {region}</h1>'
        f'</div>'
        f'<img src="{LOGO_NWC}" alt="NWC">'
        f'</div>'
        f'<div class="pmgr">{mgr_html}</div>'
    )

    cards = ''
    for re_name in re_names:
        info      = re_info.get(re_name, {'phone': '', 'office_code': '', 'office_name': re_name})
        projects  = proj_re.get(re_name, [])
        employees = emp_re.get(re_name, [])
        cards    += card_html(re_name, info, projects, employees)

    spec = spec_area_html(re_names, region_specs)

    sig_html = _build_sig_block(signatories)

    ftr = (
        f'<div class="pf">'
        f'<span>شركة مجموعة العمرو للاستشارات الهندسية</span>'
        f'<span>شركة المياه الوطنية — المنطقة الجنوبية</span>'
        f'<span>صفحة {page_num} من {total_pages}&nbsp;&nbsp;|&nbsp;&nbsp;{REPORT_MONTH}</span>'
        f'</div>'
    )

    return (
        f'<div class="page">'
        f'{hdr}'
        f'<div class="cards">{cards}</div>'
        f'{spec}'
        f'{sig_html}'
        f'{ftr}'
        f'</div>'
    )

# ── 10. Generate all regions ──────────────────────────────────────────────────
def gen_region(region):
    re_list      = reg_order.get(region, [])
    all_re_names = [rn for _, rn in re_list]
    proj_re      = proj_by_re.get(region, {})
    region_specs = spec_by_region.get(region, [])

    def _mi(role, name):
        return f'<div class="mi"><span class="ml">{role}</span>{name}</div>'
    parts = [_mi(r, n) for r, n in fixed_stakeholders]
    if region in region_stakeholder:
        parts.append(_mi(*region_stakeholder[region]))
    mgr_html = '<div class="ms"></div>'.join(parts)

    page_no_vals = [re_info.get(rn, {}).get('page_no', '') for rn in all_re_names]
    if any(page_no_vals):
        seen, grp = [], {}
        for i, pno in enumerate(page_no_vals):
            key = pno or '1'
            if key not in grp:
                grp[key] = []; seen.append(key)
            grp[key].append(i)
        page_indices = [grp[k] for k in seen]
    else:
        groups = REGION_GROUPS.get(region, [3])
        page_indices, idx = [], 0
        for n in groups:
            chunk = list(range(idx, min(idx + n, len(all_re_names))))
            if chunk: page_indices.append(chunk)
            idx += n
        if idx < len(all_re_names):
            page_indices.append(list(range(idx, len(all_re_names))))

    signatories = list(fixed_stakeholders)
    if region in region_stakeholder:
        signatories.append(region_stakeholder[region])

    total = len(page_indices)
    pages_html = ''
    for pg_num, indices in enumerate(page_indices, 1):
        re_names    = [all_re_names[i] for i in indices]
        pages_html += build_page(region, pg_num, total, re_names,
                                 proj_re, emp_by_re, mgr_html, region_specs, signatories)

    field_emp_count = sum(len(emp_by_re.get(rn, [])) for rn in all_re_names)
    tech_sup_count  = max(0, region_emp_total.get(region, 0) - field_emp_count)
    cover = build_cover(region, all_re_names, proj_re, emp_by_re,
                        mgr_html, signatories, tech_sup_count)

    html = f'''<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=1481">
<title>الهيكل التنظيمي — {region}</title>
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;900&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>
{cover}
{pages_html}
</body>
</html>'''

    fname = REGION_FILE[region]
    (OUT_DIR / fname).write_text(html, encoding='utf-8')
    print(f'  [OK] {fname}  ({len(re_list)} RE offices, {total} pages + cover)')

ALL_REGIONS = ['عسير', 'جازان', 'الباحة', 'نجران']
REGION_ARG  = {'asir': 'عسير', 'jizan': 'جازان', 'baha': 'الباحة', 'najran': 'نجران',
               'عسير': 'عسير', 'جازان': 'جازان', 'الباحة': 'الباحة', 'نجران': 'نجران'}

if __name__ == '__main__':
    arg = sys.argv[1].strip() if len(sys.argv) > 1 else None
    regions_to_run = [REGION_ARG[arg]] if arg and arg in REGION_ARG else ALL_REGIONS

    print(f"\nGenerating org charts...")
    print(f"   Output: {OUT_DIR}")
    print(f"   Data: {DATA_DIR}")
    print()

    for region in regions_to_run:
        gen_region(region)
    print(f'\nAll done!\n')
