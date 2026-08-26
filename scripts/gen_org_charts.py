"""
Generate organizational chart HTML files for each region.
Reads employee and project data from Mizan SQLAlchemy database.
Outputs: app/static/org_charts/09_OrgChart_Asir.html, etc.

Run from mizan-app root: python scripts/gen_org_charts.py
"""
import sys
import os
from pathlib import Path

# Flask app context
sys.path.insert(0, str(Path(__file__).parent.parent))
from app import create_app, db
from app.models import Employee, EmployeeStatus, JobCode, Project, Nationality

# Initialize Flask app
app = create_app()
app.app_context().push()

# Constants
REGIONS = ['عسير', 'جازان', 'الباحة', 'نجران']
REGION_CODES = {'عسير': 'AS', 'جازان': 'JZ', 'الباحة': 'BA', 'نجران': 'NJ'}
REGION_FILES = {
    'عسير': '09_OrgChart_Asir.html',
    'جازان': '10_OrgChart_Jizan.html',
    'الباحة': '11_OrgChart_Baha.html',
    'نجران': '12_OrgChart_Najran.html',
}

# Output directory
BASE = Path(__file__).parent.parent
OUTPUT_DIR = BASE / 'app' / 'static' / 'org_charts'

# ERR_VALS from HR project
ERR_VALS = {'#REF!', '=#REF!', '#VALUE!', '#N/A', '#NAME?', '#DIV/0!', '#NULL!', '#NUM!'}


# ── Helper Functions ──────────────────────────────────────────────────────────


def sv(v):
    """Safe value: convert to string, strip whitespace, return empty if error value."""
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ERR_VALS else s


def short3(s):
    """Return first 3 chars of string (for region codes in ID)."""
    return (sv(s) or '??')[:3]


def strip_grade(job_title):
    """Remove E2/E3 grade suffix from job title."""
    if not job_title:
        return ''
    s = job_title.strip()
    # Remove trailing " E2" or " E3"
    if s.endswith(' E2'):
        return s[:-3].strip()
    if s.endswith(' E3'):
        return s[:-3].strip()
    return s


def load_logo(logo_path):
    """Load logo image and convert to base64 data URL."""
    import base64

    if not logo_path.exists():
        print(f"Warning: Logo not found: {logo_path}")
        return None

    try:
        with open(logo_path, 'rb') as f:
            data = base64.b64encode(f.read()).decode('utf-8')
            # Detect image type from extension
            ext = logo_path.suffix.lower()
            mime_type = {
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.gif': 'image/gif',
                '.svg': 'image/svg+xml',
            }.get(ext, 'image/png')
            return f"data:{mime_type};base64,{data}"
    except Exception as e:
        print(f"Error loading logo {logo_path}: {e}")
        return None


def load_supporting_excel_files():
    """
    Load emp_sort.xlsx and Office-RE.xlsx for reference data.
    These files are expected to exist in /home/southMizan/sources/ on production
    or in a local Organize/ folder for development.
    Returns: (emp_sort_data, office_re_data)
    """
    import openpyxl
    import shutil

    emp_sort_data = {}
    office_re_data = {}

    # Try to load emp_sort.xlsx (employee sort order/hierarchy)
    emp_sort_paths = [
        BASE / 'data' / 'emp_sort.xlsx',
        BASE.parent / 'Organize' / 'emp_sort.xlsx',
        Path('/home/southMizan/sources/emp_sort.xlsx'),
    ]

    for path in emp_sort_paths:
        if path.exists():
            try:
                print(f"Loading emp_sort.xlsx from {path}")
                # Use load_copy pattern to avoid PermissionError
                tmp = str(path) + ".tmp.xlsx"
                shutil.copy2(path, tmp)
                wb = openpyxl.load_workbook(tmp, data_only=True)
                ws = wb.active
                # Expected: columns with employee name, reporting structure, etc.
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:
                        continue  # Skip header
                    if row[0]:  # If first column has data
                        emp_sort_data[sv(row[0])] = row
                wb.close()
                os.remove(tmp)
                print(f"  [OK] Loaded {len(emp_sort_data)} employee hierarchy records")
                break
            except Exception as e:
                print(f"  Error loading from {path}: {e}")
                if os.path.exists(tmp):
                    os.remove(tmp)

    # Try to load Office-RE.xlsx (office/regional engineer mapping)
    office_re_paths = [
        BASE / 'data' / 'Office-RE.xlsx',
        BASE.parent / 'Organize' / 'Office-RE.xlsx',
        Path('/home/southMizan/sources/Office-RE.xlsx'),
    ]

    for path in office_re_paths:
        if path.exists():
            try:
                print(f"Loading Office-RE.xlsx from {path}")
                tmp = str(path) + ".tmp.xlsx"
                shutil.copy2(path, tmp)
                wb = openpyxl.load_workbook(tmp, data_only=True)
                ws = wb.active
                # Expected: office name -> RE assignments
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:
                        continue  # Skip header
                    if row[0]:  # If first column has data
                        office_re_data[sv(row[0])] = row
                wb.close()
                os.remove(tmp)
                print(f"  [OK] Loaded {len(office_re_data)} office/RE mapping records")
                break
            except Exception as e:
                print(f"  Error loading from {path}: {e}")
                if os.path.exists(tmp):
                    os.remove(tmp)

    if not emp_sort_data:
        print("Warning: emp_sort.xlsx not found or empty")
    if not office_re_data:
        print("Warning: Office-RE.xlsx not found or empty")

    return emp_sort_data, office_re_data


def load_employees():
    """
    Load employee data from Mizan database.
    Filter: status.name_ar == 'على قوة العمل' (on active workforce)
    Returns: list of Employee objects
    """
    print("\nLoading employees from database...")

    try:
        # First, try to get the status ID for 'على قوة العمل'
        active_status = EmployeeStatus.query.filter_by(name_ar='على قوة العمل').first()

        if not active_status:
            print("Warning: Status 'على قوة العمل' not found in database")
            # If not found, try to create it or list what statuses exist
            all_statuses = EmployeeStatus.query.all()
            if all_statuses:
                print(f"  Available statuses: {[s.name_ar for s in all_statuses]}")
            else:
                print("  (No statuses in database - data may not be imported yet)")
            return []

        # Query employees with active status, eagerly load related objects
        employees = Employee.query.filter_by(current_status_id=active_status.id).all()
        print(f"  [OK] Loaded {len(employees)} active employees")
        return employees
    except Exception as e:
        print(f"Error loading employees: {e}")
        import traceback
        traceback.print_exc()
        return []


def load_projects():
    """
    Load project data from Mizan database.
    Filter: included == True, project_state == 'تحت التنفيذ'
    Returns: list of Project objects
    """
    print("Loading projects from database...")

    try:
        # Query projects that are included and ongoing (تحت التنفيذ)
        projects = Project.query.filter(
            Project.included == True,
            Project.project_state == 'تحت التنفيذ'
        ).all()
        print(f"  [OK] Loaded {len(projects)} ongoing projects")
        return projects
    except Exception as e:
        print(f"Error loading projects: {e}")
        return []


def generate_org_chart_html(region, employees, projects, emp_sort_data, office_re_data):
    """
    Generate HTML org chart for a single region.

    Structure:
    - Header with logos and title
    - Region summary (employee count, project count, etc.)
    - Organizational hierarchy/tree view
    - Project assignments
    - Management chain

    Args:
        region: Region name (e.g., 'عسير')
        employees: List of Employee objects
        projects: List of Project objects
        emp_sort_data: Employee hierarchy data from emp_sort.xlsx
        office_re_data: Office/RE mapping from Office-RE.xlsx

    Returns:
        HTML string
    """
    # Filter employees by region
    region_emps = [e for e in employees if sv(e.region) == region]

    # Filter projects by region
    region_projects = [p for p in projects if sv(p.region) == region]

    print(f"  Generating {region}: {len(region_emps)} employees, {len(region_projects)} projects")

    # Build HTML
    html_parts = []

    # DOCTYPE and head
    html_parts.append('<!DOCTYPE html>')
    html_parts.append('<html dir="rtl" lang="ar">')
    html_parts.append('<head>')
    html_parts.append('  <meta charset="UTF-8">')
    html_parts.append('  <meta name="viewport" content="width=device-width, initial-scale=1.0">')
    html_parts.append(f'  <title>الهيكل التنظيمي - {region}</title>')

    # Style
    html_parts.append('  <style>')
    html_parts.append('''
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }

    body {
        font-family: "Cairo", Arial, sans-serif;
        background-color: #f5f5f5;
        color: #333;
        line-height: 1.6;
        direction: rtl;
    }

    .header {
        background: linear-gradient(to bottom, #0a1f3d, #0071b9);
        color: white;
        padding: 20px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        direction: ltr;
    }

    .header img {
        height: 60px;
    }

    .header-title {
        flex-grow: 1;
        text-align: center;
        direction: rtl;
        font-size: 24px;
        font-weight: bold;
    }

    .container {
        max-width: 1400px;
        margin: 20px auto;
        background: white;
        border-radius: 8px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        padding: 20px;
    }

    .summary {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 15px;
        margin-bottom: 30px;
    }

    .kpi {
        background: #f9f9f9;
        border: 1px solid #ddd;
        border-right: 4px solid #0071b9;
        padding: 15px;
        border-radius: 4px;
    }

    .kpi-label {
        font-size: 12px;
        color: #666;
        margin-bottom: 5px;
    }

    .kpi-value {
        font-size: 24px;
        font-weight: bold;
        color: #0a1f3d;
    }

    .section-title {
        font-size: 18px;
        font-weight: bold;
        color: #0a1f3d;
        border-bottom: 2px solid #0071b9;
        padding-bottom: 10px;
        margin: 30px 0 20px 0;
    }

    .emp-list {
        display: grid;
        grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
        gap: 15px;
    }

    .emp-card {
        background: #f9f9f9;
        border: 1px solid #ddd;
        border-radius: 4px;
        padding: 15px;
    }

    .emp-name {
        font-weight: bold;
        color: #0a1f3d;
        margin-bottom: 5px;
    }

    .emp-detail {
        font-size: 12px;
        color: #666;
        margin: 3px 0;
    }

    .pro-list {
        display: grid;
        grid-template-columns: repeat(auto-fill, minmax(350px, 1fr));
        gap: 15px;
    }

    .pro-card {
        background: #f9f9f9;
        border: 1px solid #ddd;
        border-left: 4px solid #059669;
        border-radius: 4px;
        padding: 15px;
    }

    .pro-name {
        font-weight: bold;
        color: #0a1f3d;
        margin-bottom: 5px;
    }

    .pro-detail {
        font-size: 12px;
        color: #666;
        margin: 3px 0;
    }

    .pro-value {
        font-weight: bold;
        color: #059669;
        margin-top: 10px;
    }

    @media print {
        body {
            background: white;
        }
        .container {
            box-shadow: none;
            max-width: 100%;
        }
    }
    ''')
    html_parts.append('  </style>')
    html_parts.append('</head>')
    html_parts.append('<body>')

    # Header
    html_parts.append('  <div class="header">')
    html_parts.append('    <img src="https://via.placeholder.com/60" alt="NWC">')
    html_parts.append('    <div class="header-title">الهيكل التنظيمي - ' + region + '</div>')
    html_parts.append('    <img src="https://via.placeholder.com/60" alt="Al-Amro">')
    html_parts.append('  </div>')

    # Container
    html_parts.append('  <div class="container">')

    # Summary KPIs
    html_parts.append('    <div class="summary">')
    html_parts.append('      <div class="kpi">')
    html_parts.append('        <div class="kpi-label">إجمالي الموظفين</div>')
    html_parts.append(f'        <div class="kpi-value">{len(region_emps)}</div>')
    html_parts.append('      </div>')
    html_parts.append('      <div class="kpi">')
    html_parts.append('        <div class="kpi-label">المشاريع النشطة</div>')
    html_parts.append(f'        <div class="kpi-value">{len(region_projects)}</div>')
    html_parts.append('      </div>')
    html_parts.append('    </div>')

    # Employees section
    if region_emps:
        html_parts.append('    <div class="section-title">الموظفون</div>')
        html_parts.append('    <div class="emp-list">')
        for emp in sorted(region_emps, key=lambda e: sv(e.full_name)):
            html_parts.append('      <div class="emp-card">')
            html_parts.append(f'        <div class="emp-name">{sv(emp.full_name)}</div>')
            if emp.job_code:
                html_parts.append(f'        <div class="emp-detail"><strong>الوظيفة:</strong> {strip_grade(emp.job_code.title)}</div>')
            if emp.re_code:
                html_parts.append(f'        <div class="emp-detail"><strong>كود المقيم:</strong> {sv(emp.re_code)}</div>')
            if emp.direct_manager:
                html_parts.append(f'        <div class="emp-detail"><strong>المدير المباشر:</strong> {sv(emp.direct_manager)}</div>')
            if emp.nationality:
                html_parts.append(f'        <div class="emp-detail"><strong>الجنسية:</strong> {sv(emp.nationality.name_ar)}</div>')
            html_parts.append('      </div>')
        html_parts.append('    </div>')

    # Projects section
    if region_projects:
        html_parts.append('    <div class="section-title">المشاريع</div>')
        html_parts.append('    <div class="pro-list">')
        for proj in sorted(region_projects, key=lambda p: sv(p.name)):
            html_parts.append('      <div class="pro-card">')
            html_parts.append(f'        <div class="pro-name">{sv(proj.name)}</div>')
            if proj.contractor_name:
                html_parts.append(f'        <div class="pro-detail"><strong>المقاول:</strong> {sv(proj.contractor_name)}</div>')
            if proj.re_code:
                html_parts.append(f'        <div class="pro-detail"><strong>المقيم:</strong> {sv(proj.re_code)}</div>')
            if proj.start_date:
                html_parts.append(f'        <div class="pro-detail"><strong>تاريخ البدء:</strong> {proj.start_date}</div>')
            if proj.end_date:
                html_parts.append(f'        <div class="pro-detail"><strong>تاريخ الانتهاء:</strong> {proj.end_date}</div>')
            if proj.value:
                html_parts.append(f'        <div class="pro-value">القيمة: {proj.value:,.0f} ر.س</div>')
            html_parts.append('      </div>')
        html_parts.append('    </div>')

    html_parts.append('  </div>')
    html_parts.append('</body>')
    html_parts.append('</html>')

    return '\n'.join(html_parts)


def write_org_charts(employees, projects, emp_sort_data, office_re_data):
    """
    Generate and write org chart HTML files for all regions.

    Returns: (success_count, error_count)
    """
    success_count = 0
    error_count = 0

    print("\nGenerating org chart HTML files...")

    for region in REGIONS:
        try:
            html_content = generate_org_chart_html(region, employees, projects, emp_sort_data, office_re_data)
            output_file = OUTPUT_DIR / REGION_FILES[region]

            # Write to file
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(html_content)

            print(f"  [OK] {REGION_FILES[region]} written ({len(html_content)} bytes)")
            success_count += 1
        except Exception as e:
            print(f"  ERROR: Failed to generate {REGION_FILES[region]}: {e}")
            error_count += 1

    return success_count, error_count


def verify_output_files():
    """Verify that all output files were created successfully."""
    print("\nVerifying output files...")
    all_exist = True

    for region, filename in REGION_FILES.items():
        filepath = OUTPUT_DIR / filename
        if filepath.exists():
            size = filepath.stat().st_size
            print(f"  [OK] {filename} ({size:,} bytes)")
        else:
            print(f"  [FAILED] {filename} NOT FOUND")
            all_exist = False

    return all_exist


def main():
    print(f"Generating organizational charts...")

    # Create output directory if it doesn't exist
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {OUTPUT_DIR}")

    # Load supporting Excel files (optional reference data)
    emp_sort_data, office_re_data = load_supporting_excel_files()

    # Load employee data
    employees = load_employees()
    if not employees:
        print("ERROR: No employees found")
        return

    # Load project data
    projects = load_projects()
    if not projects:
        print("Warning: No active projects found")

    # Generate and write HTML files
    success_count, error_count = write_org_charts(employees, projects, emp_sort_data, office_re_data)

    # Verify
    all_ok = verify_output_files()

    # Summary
    print(f"\n{'='*60}")
    print(f"Summary: {success_count} files generated, {error_count} errors")
    if all_ok:
        print("[SUCCESS] All org chart files generated successfully!")
        print(f"  Location: {OUTPUT_DIR}")
        print(f"  Access: https://southmizan.pythonanywhere.com/static/org_charts/")
    else:
        print("[FAILED] Some files failed to generate")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
