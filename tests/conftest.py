import pytest
from openpyxl import Workbook
from app import create_app, db as _db


def _blank_emp_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'data'
    ws.append(['h'] * 28)
    wb.save(path)


def _blank_pro_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'pro'
    ws.append(['h'] * 34)
    wb.save(path)


def _blank_re_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'RE_Mail'
    ws.append(['h'] * 10)
    wb.save(path)


@pytest.fixture
def app(tmp_path):
    emp = tmp_path / 'emp.xlsx'
    pro = tmp_path / 'pro.xlsx'
    re_xlsx = tmp_path / 'office_re.xlsx'
    _blank_emp_xlsx(emp)
    _blank_pro_xlsx(pro)
    _blank_re_xlsx(re_xlsx)
    app = create_app({
        'TESTING': True,
        'SQLALCHEMY_DATABASE_URI': 'sqlite:///:memory:',
        'SECRET_KEY': 'test-secret',
        'WTF_CSRF_ENABLED': False,
        'EXCEL_SOURCES': {
            'employees': str(emp),
            'projects': str(pro),
            'office_re': str(re_xlsx),
        },
    })
    with app.app_context():
        _db.create_all()
        yield app
        _db.drop_all()


@pytest.fixture
def client(app):
    return app.test_client()
