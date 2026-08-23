import pytest
import openpyxl
import tempfile
import shutil
from pathlib import Path
from app.models import Employee, EmployeeStatus
from app import db


def test_export_writes_active_employees_only(app):
    with app.app_context():
        active = EmployeeStatus(name_ar='على قوة العمل')
        left   = EmployeeStatus(name_ar='استقالة')
        emp1   = Employee(full_name='موظف نشط',     cbu='JZ', region='جازان', current_status=active)
        emp2   = Employee(full_name='موظف مستقيل', cbu='AS', region='عسير',  current_status=left)
        db.session.add_all([active, left, emp1, emp2])
        db.session.commit()

    # Create a minimal Excel for testing
    tmp_dir = tempfile.mkdtemp()
    fake_excel = Path(tmp_dir) / 'test_emp.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'data'
    ws.append([None] * 32)  # header row with 32 empty cells
    wb.save(str(fake_excel))

    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from scripts.export_emp_excel import export_to_excel

    with app.app_context():
        count = export_to_excel(excel_path=str(fake_excel))

    shutil.rmtree(tmp_dir)
    assert count == 1  # only active employee


def test_export_name_in_correct_column(app):
    with app.app_context():
        active = EmployeeStatus(name_ar='على قوة العمل')
        emp = Employee(full_name='فهد العمري', cbu='NJ', region='نجران', current_status=active)
        db.session.add_all([active, emp])
        db.session.commit()

    tmp_dir = tempfile.mkdtemp()
    fake_excel = Path(tmp_dir) / 'test_emp2.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'data'
    ws.append([None] * 32)
    wb.save(str(fake_excel))

    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from scripts.export_emp_excel import export_to_excel

    with app.app_context():
        export_to_excel(excel_path=str(fake_excel))

    # Re-read and verify name is at column 21 (0-based index 20 → openpyxl col 21)
    wb2 = openpyxl.load_workbook(str(fake_excel))
    ws2 = wb2['data']
    name_cell = ws2.cell(row=2, column=21).value
    shutil.rmtree(tmp_dir)
    assert name_cell == 'فهد العمري'


def test_export_status_in_correct_column(app):
    with app.app_context():
        active = EmployeeStatus(name_ar='على قوة العمل')
        emp = Employee(full_name='نورة السالم', cbu='BA', region='الباحة', current_status=active)
        db.session.add_all([active, emp])
        db.session.commit()

    tmp_dir = tempfile.mkdtemp()
    fake_excel = Path(tmp_dir) / 'test_status.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'data'
    ws.append([None] * 32)
    wb.save(str(fake_excel))

    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from scripts.export_emp_excel import export_to_excel

    with app.app_context():
        export_to_excel(excel_path=str(fake_excel))

    wb2 = openpyxl.load_workbook(str(fake_excel))
    ws2 = wb2['data']
    # col index 11 (0-based) → openpyxl col 12
    status_cell = ws2.cell(row=2, column=12).value
    shutil.rmtree(tmp_dir)
    assert status_cell == 'على قوة العمل'


def test_export_backup_created(app):
    with app.app_context():
        active = EmployeeStatus(name_ar='على قوة العمل')
        emp = Employee(full_name='ماجد القحطاني', cbu='AS', region='عسير', current_status=active)
        db.session.add_all([active, emp])
        db.session.commit()

    tmp_dir = tempfile.mkdtemp()
    fake_excel = Path(tmp_dir) / 'test_backup.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'data'
    ws.append([None] * 32)
    wb.save(str(fake_excel))

    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from scripts.export_emp_excel import export_to_excel

    with app.app_context():
        export_to_excel(excel_path=str(fake_excel))

    # Check that a .bak_ file was created alongside the excel
    bak_files = list(Path(tmp_dir).glob('*.bak_*'))
    shutil.rmtree(tmp_dir)
    assert len(bak_files) == 1, f'Expected 1 backup file, found: {bak_files}'
