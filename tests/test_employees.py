from pathlib import Path
from openpyxl import load_workbook
from app.models import Employee, EmployeeStatus, EmployeeStatusHistory, User
from app import db

# ── helpers ────────────────────────────────────────────────────────────────────

def _login(client, app):
    with app.app_context():
        u = User(username='op', role='admin')
        u.set_password('pw')
        db.session.add(u)
        db.session.commit()
    client.post('/auth/login', data={'username': 'op', 'password': 'pw'})

def _seed_employee(app, name='سعد محمد', region='جازان', status='على قوة العمل'):
    path = Path(app.config['EXCEL_SOURCES']['employees'])
    wb = load_workbook(path)
    ws = wb['data']
    row_idx = ws.max_row + 1
    ws.cell(row_idx, 12, status)   # col 11 status
    ws.cell(row_idx, 17, region)   # col 16 region
    ws.cell(row_idx, 21, name)     # col 20 name
    ws.cell(row_idx, 22, 'مهندس موقع E3')
    wb.save(path)
    return row_idx

# ── list tests ─────────────────────────────────────────────────────────────────

def test_create_employee(app):
    with app.app_context():
        status = EmployeeStatus(name_ar='على قوة العمل')
        emp = Employee(full_name='محمد علي', cbu='JZ', region='جازان', current_status=status)
        db.session.add_all([status, emp])
        db.session.commit()
        assert emp.id is not None

def test_status_history(app):
    with app.app_context():
        from datetime import date
        status = EmployeeStatus(name_ar='على قوة العمل')
        emp = Employee(full_name='أحمد محمد', cbu='AS', region='عسير', current_status=status)
        db.session.add_all([status, emp])
        db.session.flush()
        hist = EmployeeStatusHistory(
            employee_id=emp.id, old_status=None, new_status='على قوة العمل',
            change_date=date.today(), reason='استيراد أولي',
        )
        db.session.add(hist)
        db.session.commit()
        assert emp.status_history.count() == 1

def test_employee_list_requires_login(client):
    r = client.get('/employees/', follow_redirects=False)
    assert r.status_code == 302
    assert '/auth/login' in r.headers['Location']

def test_employee_list_shows_employee(app, client):
    _login(client, app)
    _seed_employee(app)
    r = client.get('/employees/')
    assert r.status_code == 200
    assert 'سعد محمد' in r.data.decode('utf-8')

def test_employee_list_filter_by_region(app, client):
    _login(client, app)
    _seed_employee(app, name='موظف جازان', region='جازان')
    _seed_employee(app, name='موظف عسير', region='عسير')
    r = client.get('/employees/?region=جازان')
    body = r.data.decode('utf-8')
    assert 'موظف جازان' in body
    assert 'موظف عسير' not in body

def test_side_panel_returns_employee_name(app, client):
    _login(client, app)
    eid = _seed_employee(app, name='فهد العتيبي')
    r = client.get(f'/employees/{eid}/panel')
    assert r.status_code == 200
    assert 'فهد العتيبي' in r.data.decode('utf-8')

def test_profile_shows_5_sections(app, client):
    _login(client, app)
    eid = _seed_employee(app)
    r = client.get(f'/employees/{eid}')
    body = r.data.decode('utf-8')
    assert 'الهوية والتواصل' in body
    assert 'العقد والراتب' in body
    assert 'التصنيف والموقع' in body
    assert 'الاستبدال' in body
    assert 'سجل الحالة' in body

def test_status_change_is_blocked_excel_phase(app, client):
    _login(client, app)
    eid = _seed_employee(app)
    r = client.post(f'/employees/{eid}/status', data={
        'new_status_id': 1,
        'reason': 'قدم استقالته',
        'change_date': '2026-08-15',
    }, follow_redirects=True)
    assert r.status_code == 403


def test_add_employee_is_blocked_excel_phase(app, client):
    _login(client, app)
    r = client.post('/employees/add', data={'full_name': 'خالد إبراهيم عمر'}, follow_redirects=True)
    assert r.status_code == 403


def test_edit_employee_is_blocked_excel_phase(app, client):
    _login(client, app)
    eid = _seed_employee(app, name='الاسم القديم')
    r = client.post(f'/employees/{eid}/edit', data={'full_name': 'الاسم الجديد'}, follow_redirects=True)
    assert r.status_code == 403


def _login_viewer(client, app):
    with app.app_context():
        u = User(username='view', role='viewer')
        u.set_password('pw')
        db.session.add(u)
        db.session.commit()
    client.post('/auth/login', data={'username': 'view', 'password': 'pw'})


def test_viewer_cannot_change_status(app, client):
    _login_viewer(client, app)
    eid = _seed_employee(app)
    with app.app_context():
        new_st = EmployeeStatus(name_ar='استقالة')
        db.session.add(new_st)
        db.session.commit()
        new_st_id = new_st.id
    r = client.post(f'/employees/{eid}/status', data={
        'new_status_id': new_st_id,
        'reason': 'should be blocked',
        'change_date': '2026-08-15',
    })
    assert r.status_code == 403
    with app.app_context():
        assert EmployeeStatusHistory.query.filter_by(employee_id=eid).count() == 0


def test_viewer_cannot_edit_or_add(app, client):
    _login_viewer(client, app)
    eid = _seed_employee(app)
    assert client.get(f'/employees/{eid}/edit').status_code == 403
    assert client.get('/employees/add').status_code == 403


def test_viewer_profile_hides_write_actions(app, client):
    _login_viewer(client, app)
    eid = _seed_employee(app)
    body = client.get(f'/employees/{eid}').data.decode('utf-8')
    assert 'تغيير الحالة' not in body
    assert f'/employees/{eid}/edit' not in body


def test_side_panel_hides_edit_for_viewer(app, client):
    _login_viewer(client, app)
    eid = _seed_employee(app, name='فهد العتيبي')
    body = client.get(f'/employees/{eid}/panel').data.decode('utf-8')
    assert 'فهد العتيبي' in body
    assert 'تعديل' not in body


def test_region_filter_preserves_search(app, client):
    _login(client, app)
    _seed_employee(app)
    body = client.get('/employees/?q=سعد').data.decode('utf-8')
    assert '/employees/?region=عسير&status=active&q=سعد' in body
    assert '/employees/?status=active&q=سعد' in body


def test_status_change_to_replacement_is_blocked(app, client):
    _login(client, app)
    eid = _seed_employee(app)
    r = client.post(f'/employees/{eid}/status', data={
        'new_status_id': 1,
        'reason': 'تعيين بديل',
        'change_date': '2026-08-15',
    })
    assert r.status_code == 403


def test_status_change_from_replacement_is_blocked(app, client):
    _login(client, app)
    eid = _seed_employee(app, status='بديل')
    r = client.post(f'/employees/{eid}/status', data={
        'new_status_id': 1,
        'reason': 'انتهاء البديل',
        'change_date': '2026-08-15',
    })
    assert r.status_code == 403


def test_profile_salary_label_is_monthly(app, client):
    _login(client, app)
    eid = _seed_employee(app)
    body = client.get(f'/employees/{eid}').data.decode('utf-8')
    assert 'الراتب الشهري' in body
    assert 'الراتب اليومي' not in body
