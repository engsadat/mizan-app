from tests.test_dashboard_cards import _login
from openpyxl import load_workbook
from pathlib import Path


def test_reports_index(app, client):
    _login(client, app, 'ed1', 'editor')
    r = client.get('/reports/')
    body = r.data.decode('utf-8')
    assert r.status_code == 200
    assert 'لوحة إحصاءات' in body
    assert 'قابل للطباعة' in body
    assert '/reports/emp-dashboard' in body
    assert '/reports/filter-report' in body
    assert '/reports/projects-dashboard' in body


def test_projects_dashboard_reads_excel_value(app, client):
    _login(client, app, 'ed1', 'editor')
    path = Path(app.config['EXCEL_SOURCES']['projects'])
    wb = load_workbook(path)
    ws = wb['pro']
    r = ws.max_row + 1
    ws.cell(r, 11, 'yes')
    ws.cell(r, 13, 'مشروع اختبار')
    ws.cell(r, 18, 'عسير')
    ws.cell(r, 24, 'تحت التنفيذ')
    ws.cell(r, 30, 1_000_000)
    wb.save(path)
    body = client.get('/reports/projects-dashboard').data.decode('utf-8')
    assert '1.0' in body
    assert 'تحت التنفيذ' in body


def test_employee_excel_download(app, client):
    from tests.test_employees import _login as emp_login, _seed_employee
    emp_login(client, app)
    _seed_employee(app)
    r = client.get('/employees/export.xlsx')
    assert r.status_code == 200
    assert 'spreadsheet' in (r.mimetype or '') or r.data[:2] == b'PK'


def test_norm_person_collapses_yeh_and_spaces():
    from app.excel_data import norm_person, match_office_name
    assert norm_person('محمد مصطفى  محمود') == norm_person('محمد مصطفي محمود')
    canon = {norm_person('محمد مصطفى محمود'): 'محمد مصطفى محمود'}
    assert match_office_name('محمد مصطفي محمود', canon) == 'محمد مصطفى محمود'
    assert match_office_name('محمد مصطفى  محمود', canon) == 'محمد مصطفى محمود'


def test_org_smart_search_includes_job_and_office(app, client):
    from openpyxl import load_workbook
    _login(client, app, 'ed1', 'editor')
    re_path = Path(app.config['EXCEL_SOURCES']['office_re'])
    wb = load_workbook(re_path)
    ws = wb['RE_Mail']
    r = ws.max_row + 1
    ws.cell(r, 2, 'محمود عطيتو محمد محمود')
    ws.cell(r, 5, 'عسير')
    ws.cell(r, 9, 'أبها_1')
    wb.save(re_path)

    emp_path = Path(app.config['EXCEL_SOURCES']['employees'])
    wb = load_workbook(emp_path)
    ws = wb['data']
    r = ws.max_row + 1
    ws.cell(r, 12, 'على قوة العمل')
    ws.cell(r, 17, 'عسير')
    ws.cell(r, 21, 'سعد الاختبار')
    ws.cell(r, 22, 'مهندس موقع E3')
    ws.cell(r, 28, 'محمود عطيتو  محمد محمود')
    wb.save(emp_path)

    body = client.get('/reports/org-chart-smart').data.decode('utf-8')
    assert 'أبها_1' in body
    assert 'مهندس موقع' in body
    assert 'data-searchtext' in body
    assert 'مهندس' in body


def test_org_chart_print_is_tel_version(app, client):
    _login(client, app, 'ed1', 'editor')
    r = client.get('/reports/org-chart/asir')
    body = r.data.decode('utf-8')
    assert r.status_code == 200
    assert 'مع الاتصال' in body
    assert 'class="ep"' in body
    assert 'المنطقة الجنوبية - NWC' not in body
