"""Safe Excel read/write utilities with backup and error handling."""

import os
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


ERR_VALS = {'#REF!', '=#REF!', '#VALUE!', '#N/A', '#NAME?', '#DIV/0!', '#NULL!', '#NUM!'}


def sv(v):
    """Safe value: strip and return empty string if Excel error."""
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ERR_VALS else s


def load_copy(path):
    """
    Load Excel file safely by copying to temp file first.
    Avoids PermissionError if file is open in Excel.

    Args:
        path: Path to Excel file

    Returns:
        openpyxl Workbook object

    Raises:
        FileNotFoundError: If file doesn't exist
        openpyxl.exc.InvalidFileException: If not a valid Excel file
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    tmp = path.parent / (path.stem + '.tmp' + path.suffix)
    try:
        shutil.copy2(path, tmp)
        wb = load_workbook(tmp, data_only=True)
        return wb
    finally:
        if tmp.exists():
            try:
                os.remove(tmp)
            except OSError:
                pass  # Already deleted or locked, ignore


def save_excel(wb, path, backup=True):
    """
    Save Excel file with optional timestamped backup.

    Args:
        wb: openpyxl Workbook object
        path: Path to save
        backup: If True, save backup as path.backup.YYYYMMDD_HHMMSS.xlsx

    Raises:
        IOError: If save fails
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if backup and path.exists():
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = path.with_stem(f"{path.stem}.backup.{timestamp}")
        shutil.copy2(path, backup_path)

    wb.save(path)


def read_sheet(path, sheet_name=None, min_row=2):
    """
    Read Excel sheet as list of dicts (header in row 1).

    Args:
        path: Path to Excel file
        sheet_name: Sheet name (default: first sheet)
        min_row: Skip rows before this (default: 2, to skip header)

    Returns:
        List of dicts {col_name: value, ...}
    """
    wb = load_copy(path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if ws.max_row < 2:
        return []

    headers = []
    for cell in ws[1]:
        headers.append(sv(cell.value))

    rows = []
    for row in ws.iter_rows(min_row=min_row, values_only=False):
        row_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = sv(cell.value)
        rows.append(row_dict)

    wb.close()
    return rows


def read_column(path, sheet_name, column_idx, min_row=2):
    """
    Read single column as list.

    Args:
        path: Path to Excel file
        sheet_name: Sheet name
        column_idx: 0-based column index
        min_row: Skip rows before this

    Returns:
        List of values
    """
    wb = load_copy(path)
    ws = wb[sheet_name]

    values = []
    for row in ws.iter_rows(min_row=min_row, max_col=column_idx + 1, values_only=False):
        if row and len(row) > column_idx:
            values.append(sv(row[column_idx].value))

    wb.close()
    return values
